import customtkinter as ctk
import tkinter as tk
from config.themes import COLORS, FONTS
from ui.components.data_table import DataTable
from ui.components.responsive import Debouncer
from ui.components.toast import Toast
from database.engine import get_session
from core.services.report_service import ReportService
from core.services.audit_service import AuditService
import os
import subprocess
import sys
from datetime import date


class ReportsScreen(ctk.CTkFrame):

    def __init__(self, parent, app, **kwargs):
        super().__init__(
            parent, fg_color=COLORS["bg"], corner_radius=0, **kwargs)
        self.app = app
        self._active_tab = "inventory"
        self._inv_low_only = False   # Parts Inventory: show only low-stock parts
        self._report_date_var = tk.StringVar(value=date.today().isoformat())
        self._build()

    def _build(self):
        # ── Top bar ───────────────────────────────────────────────────
        topbar = ctk.CTkFrame(self, fg_color=COLORS["card"],
                              corner_radius=0, height=60)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        ctk.CTkLabel(topbar, text="Reports",
                     font=FONTS["title"],
                     text_color=COLORS["txt"]).pack(side="left", padx=24, pady=16)

        ctk.CTkButton(topbar, text="⬇  Export Excel",
                      fg_color=COLORS["navy"], hover_color=COLORS["navy_hover"],
                      text_color="#FFFFFF", font=FONTS["body"],
                      width=120, height=34,
                      command=self._export).pack(side="right", padx=16, pady=13)

        ctk.CTkButton(topbar, text="Apply",
                      fg_color=COLORS["bg2"], hover_color=COLORS["border"],
                      text_color=COLORS["txt2"], font=FONTS["small"],
                      width=58, height=32,
                      command=self.refresh).pack(side="right", padx=(4, 0), pady=14)

        ctk.CTkButton(topbar, text="Today",
                      fg_color=COLORS["bg2"], hover_color=COLORS["border"],
                      text_color=COLORS["txt2"], font=FONTS["small"],
                      width=60, height=32,
                      command=self._set_today).pack(side="right", padx=(6, 0), pady=14)

        ctk.CTkButton(topbar, text="All",
                      fg_color=COLORS["bg2"], hover_color=COLORS["border"],
                      text_color=COLORS["txt2"], font=FONTS["small"],
                      width=45, height=32,
                      command=self._set_all_dates).pack(side="right", padx=(6, 0), pady=14)

        ctk.CTkEntry(topbar, textvariable=self._report_date_var,
                     placeholder_text="YYYY-MM-DD",
                     fg_color=COLORS["bg2"], border_color=COLORS["border"],
                     text_color=COLORS["txt"], font=FONTS["small"],
                     width=112, height=32).pack(side="right", padx=(8, 0), pady=14)

        ctk.CTkLabel(topbar, text="Report Date",
                     font=FONTS["small"],
                     text_color=COLORS["txt3"]).pack(side="right", padx=(0, 4), pady=14)

        # ── Tab bar ───────────────────────────────────────────────────
        tab_bar = ctk.CTkFrame(self, fg_color=COLORS["card"],
                               corner_radius=0, height=46)
        tab_bar.pack(fill="x")
        ctk.CTkFrame(self, fg_color=COLORS["border"], height=1).pack(fill="x")

        self._tab_btns = {}
        tabs = [
            ("Parts Inventory",  "inventory"),
            ("By Category",      "category"),
            ("Movement History", "movements"),
            ("Audit Log",        "audit"),
            ("Adjustments",      "adjustments"),
            ("Aging Analysis",   "aging"),
            ("Sales",            "sales"),
            ("Returns",          "returns"),
        ]
        inner = ctk.CTkFrame(tab_bar, fg_color="transparent")
        inner.pack(side="left", padx=16, pady=6)

        for label, key in tabs:
            btn = ctk.CTkButton(
                inner, text=label,
                fg_color=COLORS["navy"] if key == self._active_tab else COLORS["bg2"],
                hover_color=COLORS["navy_hover"] if key == self._active_tab else COLORS["border"],
                text_color="#FFFFFF" if key == self._active_tab else COLORS["txt2"],
                font=FONTS["body"],
                height=32, width=118, corner_radius=8,
                command=lambda k=key: self._switch_tab(k),
            )
            btn.pack(side="left", padx=(0, 6))
            self._tab_btns[key] = btn

        # ── Content area ──────────────────────────────────────────────
        self._content = ctk.CTkFrame(
            self, fg_color=COLORS["bg"], corner_radius=0)
        self._content.pack(fill="both", expand=True)

        self._switch_tab("inventory")

    def refresh(self):
        self._switch_tab(self._active_tab)

    def _set_today(self):
        self._report_date_var.set(date.today().isoformat())
        self.refresh()

    def _set_all_dates(self):
        self._report_date_var.set("")
        self.refresh()

    def _selected_date(self):
        """
        Reports are filtered by one report date.
        Default is today. If the field is cleared, reports show all dates.
        """
        value = self._report_date_var.get().strip()
        if not value:
            return None
        try:
            # Validate YYYY-MM-DD format
            date.fromisoformat(value)
            return value
        except ValueError:
            Toast(self.app, "Use date format YYYY-MM-DD", kind="error")
            return None

    def _date_kwargs(self):
        d = self._selected_date()
        return {"date_from": d, "date_to": d} if d else {}

    def _inventory_matches_date(self, row, selected_date):
        if not selected_date:
            return True
        return row.get("last_received") == selected_date or row.get("last_issued") == selected_date

    def _filter_inventory_for_date(self, rows):
        selected_date = self._selected_date()
        if not selected_date:
            return rows
        return [r for r in rows if self._inventory_matches_date(r, selected_date)]

    def _summary_from_inventory(self, rows):
        return {
            "total_parts": len(rows),
            "total_units": sum(r.get("current_stock", 0) or 0 for r in rows),
            "total_value": round(sum(r.get("stock_value", 0) or 0 for r in rows), 2),
            "low_stock": sum(1 for r in rows if r.get("is_low_stock")),
            "zero_stock": sum(1 for r in rows if (r.get("current_stock", 0) or 0) <= 0),
        }

    def _category_from_inventory(self, rows):
        grouped = {}
        for r in rows:
            name = r.get("category") or "Uncategorized"
            g = grouped.setdefault(name, {
                "category": name,
                "parts": 0,
                "total_stock": 0,
                "value": 0,
                "low_count": 0,
            })
            g["parts"] += 1
            g["total_stock"] += r.get("current_stock", 0) or 0
            g["value"] += r.get("stock_value", 0) or 0
            if r.get("is_low_stock"):
                g["low_count"] += 1
        return sorted(grouped.values(), key=lambda x: x["value"], reverse=True)

    # ── Tab switching ─────────────────────────────────────────────────
    def _switch_tab(self, key: str):
        self._active_tab = key
        for k, btn in self._tab_btns.items():
            active = k == key
            btn.configure(
                fg_color=COLORS["navy"] if active else COLORS["bg2"],
                hover_color=COLORS["navy_hover"] if active else COLORS["border"],
                text_color="#FFFFFF" if active else COLORS["txt2"],
            )
        for w in self._content.winfo_children():
            w.destroy()

        dispatch = {
            "inventory":   self._render_inventory,
            "category":    self._render_category,
            "movements":   self._render_movements,
            "audit":       self._render_audit,
            "adjustments": self._render_adjustments,
            "aging":       self._render_aging,
            "sales":       self._render_sales,
            "returns":     self._render_returns,
        }
        dispatch.get(key, lambda: None)()

    # ── Summary stat cards row ────────────────────────────────────────
    def _stat_cards(self, parent, stats: list):
        """stats = list of (label, value, accent_color)"""
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=16, pady=(12, 8))
        for i in range(len(stats)):
            row.columnconfigure(i, weight=1, uniform="stat")

        for i, (label, value, color) in enumerate(stats):
            card = ctk.CTkFrame(row, fg_color=COLORS["card"],
                                corner_radius=10,
                                border_width=1,
                                border_color=COLORS["border"])
            card.grid(row=0, column=i,
                      padx=(0 if i == 0 else 6, 0),
                      sticky="nsew", ipady=2)
            ctk.CTkLabel(card, text=label.upper(),
                         font=FONTS["label"],
                         text_color=COLORS["txt3"]).pack(anchor="w", padx=14, pady=(10, 2))
            ctk.CTkLabel(card, text=value,
                         font=("Helvetica", 18, "bold"),
                         text_color=color or COLORS["txt"]).pack(anchor="w", padx=14, pady=(0, 10))

    def _table_frame(self) -> ctk.CTkFrame:
        f = ctk.CTkFrame(
            self._content, fg_color=COLORS["card"], corner_radius=0)
        f.pack(fill="both", expand=True)
        return f

    def _toggle_inv_low(self):
        """Toggle the Parts Inventory low-stock-only filter and restyle the button."""
        self._inv_low_only = not self._inv_low_only
        active = self._inv_low_only
        self._inv_low_btn.configure(
            fg_color=COLORS["red"] if active else COLORS["bg2"],
            hover_color=COLORS["red"] if active else COLORS["border"],
            text_color="#FFFFFF" if active else COLORS["txt2"])
        if getattr(self, "_inv_filter_fn", None):
            self._inv_filter_fn()

    # ── Tab: Parts Inventory (full detail) ────────────────────────────
    def _render_inventory(self):
        db = get_session()
        try:
            all_data = ReportService(db).inventory_detail()
        finally:
            db.close()

        # Parts Inventory is a current-stock snapshot — show all parts regardless of date.
        data = all_data
        stats = self._summary_from_inventory(data)

        # Summary cards
        self._stat_cards(self._content, [
            ("Total Parts",   str(
                stats["total_parts"]),              COLORS["txt"]),
            ("Total Units",   str(
                stats["total_units"]),              COLORS["blue"]),
            ("Total Value",
             f"₱{stats['total_value']:,.2f}",        COLORS["navy"]),
            ("Low Stock",     str(stats["low_stock"]),
             COLORS["red"] if stats["low_stock"] else COLORS["green"]),
            ("Zero Stock",    str(stats["zero_stock"]),
             COLORS["red"] if stats["zero_stock"] else COLORS["green"]),
        ])

        # Search bar
        search_frame = ctk.CTkFrame(self._content, fg_color="transparent")
        search_frame.pack(fill="x", padx=16, pady=(0, 8))
        search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=search_var,
            placeholder_text="🔍  Filter by name, SKU, category, bin…",
            fg_color=COLORS["card"],
            border_color=COLORS["border"],
            text_color=COLORS["txt"],
            font=FONTS["body"],
            height=34, width=360,
        )
        search_entry.pack(side="left")

        # One-click "Low Stock only" filter beside the search bar. The Export
        # button respects it, so you can export just the low-stock list.
        self._inv_low_btn = ctk.CTkButton(
            search_frame, text="⚠  Low Stock only", font=FONTS["body"],
            fg_color=COLORS["red"] if self._inv_low_only else COLORS["bg2"],
            hover_color=COLORS["red"] if self._inv_low_only else COLORS["border"],
            text_color="#FFFFFF" if self._inv_low_only else COLORS["txt2"],
            width=152, height=34, command=self._toggle_inv_low)
        self._inv_low_btn.pack(side="left", padx=(10, 0))

        count_lbl = ctk.CTkLabel(search_frame, text=f"{len(data)} parts",
                                 font=FONTS["small"], text_color=COLORS["txt3"])
        count_lbl.pack(side="right")

        COLS = [
            {"id": "sku",      "label": "SKU",
                "width": 130, "stretch": False},
            {"id": "name",     "label": "Part Name",     "width": 200},
            {"id": "category", "label": "Category",
                "width": 110, "stretch": False},
            {"id": "stock",    "label": "Stock",         "width": 70,
                "stretch": False, "anchor": "center"},
            {"id": "unit",     "label": "Unit",          "width": 55,
                "stretch": False, "anchor": "center"},
            {"id": "min",      "label": "Min",           "width": 50,
                "stretch": False, "anchor": "center"},
            {"id": "status",   "label": "Status",        "width": 80,
                "stretch": False, "anchor": "center"},
            {"id": "cost",     "label": "Unit Cost",
                "width": 95,  "stretch": False, "anchor": "e"},
            {"id": "value",    "label": "Total Value",
                "width": 110, "stretch": False, "anchor": "e"},
            {"id": "total_in", "label": "Total In",      "width": 75,
                "stretch": False, "anchor": "center"},
            {"id": "total_out", "label": "Total Out",
                "width": 75,  "stretch": False, "anchor": "center"},
            {"id": "bin",      "label": "Bin",
                "width": 80,  "stretch": False},
            {"id": "supplier", "label": "Last Supplier",
                "width": 150, "stretch": False},
            {"id": "received", "label": "Last Received",
                "width": 110, "stretch": False},
            {"id": "issued",   "label": "Last Issued",
                "width": 110, "stretch": False},
        ]

        tbl = DataTable(self._table_frame(), COLS, height=18)
        tbl.pack(fill="both", expand=True)

        all_rows = []
        for d in data:
            tag = "low" if d["is_low_stock"] else ""
            all_rows.append({
                "id":     d["sku"],
                "_tag":   tag,
                "_data":  d,
                "values": (
                    d["sku"],
                    d["name"],
                    d["category"],
                    d["current_stock"],
                    d["unit"],
                    d["min_stock"],
                    d["status"],
                    f"₱{d['unit_cost']:,.2f}",
                    f"₱{d['stock_value']:,.2f}",
                    d["total_in"],
                    d["total_out"],
                    d["bin_location"],
                    d["last_supplier"],
                    d["last_received"],
                    d["last_issued"],
                ),
            })

        tbl.load(all_rows)

        def _filter(*_):
            term = search_var.get().strip().lower()
            rows = all_rows
            if self._inv_low_only:
                rows = [r for r in rows if r["_data"].get("is_low_stock")]
            if term:
                rows = [r for r in rows if any(
                    term in str(v).lower() for v in r["values"])]
            tbl.load(rows)
            label = "low-stock parts" if self._inv_low_only else "parts"
            count_lbl.configure(text=f"{len(rows)} {label}")

        self._inv_filter_fn = _filter
        debounce = Debouncer(self, delay_ms=160)
        search_var.trace_add("write", lambda *_: debounce.call(_filter))
        _filter()   # apply any active low-stock filter on first render

    # ── Tab: By Category ──────────────────────────────────────────────
    def _render_category(self):
        db = get_session()
        try:
            all_data = ReportService(db).inventory_detail()
        finally:
            db.close()

        # By Category is a current-stock snapshot — show all parts regardless of date.
        filtered_inventory = all_data
        data = self._category_from_inventory(filtered_inventory)
        stats = self._summary_from_inventory(filtered_inventory)

        self._stat_cards(self._content, [
            ("Total Parts",  str(stats["total_parts"]),       COLORS["txt"]),
            ("Total Value",  f"₱{stats['total_value']:,.2f}", COLORS["navy"]),
            ("Low Stock",    str(stats["low_stock"]),
             COLORS["red"] if stats["low_stock"] else COLORS["green"]),
        ])

        COLS = [
            {"id": "cat",      "label": "Category",         "width": 180},
            {"id": "parts",    "label": "Total Parts",
                "width": 100, "stretch": False, "anchor": "center"},
            {"id": "stock",    "label": "Total Stock",
                "width": 110, "stretch": False, "anchor": "center"},
            {"id": "value",    "label": "Stock Value",
                "width": 150, "stretch": False, "anchor": "e"},
            {"id": "low",      "label": "Low Stock Items",
                "width": 130, "stretch": False, "anchor": "center"},
            {"id": "pct",      "label": "% of Total Value",
                "width": 140, "stretch": False, "anchor": "e"},
        ]
        tbl = DataTable(self._table_frame(), COLS, height=20)
        tbl.pack(fill="both", expand=True)

        total_val = sum(r["value"] for r in data) or 1
        rows = []
        for i, d in enumerate(data):
            pct = (d["value"] / total_val * 100)
            rows.append({"id": i, "values": (
                d["category"],
                d["parts"],
                d["total_stock"],
                f"₱{d['value']:,.2f}",
                str(d["low_count"]) if d["low_count"] else "—",
                f"{pct:.1f}%",
            ), "_tag": "warn" if d["low_count"] > 0 else ""})

        # Totals row
        rows.append({"id": -1, "values": (
            "TOTAL",
            sum(d["parts"] for d in data),
            sum(d["total_stock"] for d in data),
            f"₱{sum(d['value'] for d in data):,.2f}",
            str(sum(d["low_count"] for d in data)),
            "100.0%",
        ), "_tag": ""})
        tbl.load(rows)

    # ── Tab: Movement History ─────────────────────────────────────────
    def _render_movements(self):
        db = get_session()
        try:
            data = ReportService(db).movement_history(**self._date_kwargs())
        finally:
            db.close()

        # Quick stats
        total_in = sum(d["quantity"] for d in data if d["type"] == "IN")
        total_out = sum(d["quantity"] for d in data if d["type"] == "OUT")
        self._stat_cards(self._content, [
            ("Total Transactions", str(len(data)),   COLORS["txt"]),
            ("Units Received",     str(total_in),    COLORS["green"]),
            ("Units Issued",       str(total_out),   COLORS["amber"]),
        ])

        # Filter bar
        fbar = ctk.CTkFrame(self._content, fg_color="transparent")
        fbar.pack(fill="x", padx=16, pady=(0, 8))
        filter_var = tk.StringVar()
        type_var = tk.StringVar(value="All")

        ctk.CTkEntry(fbar, textvariable=filter_var,
                     placeholder_text="🔍  Filter by part name, SKU, supplier…",
                     fg_color=COLORS["card"], border_color=COLORS["border"],
                     text_color=COLORS["txt"], font=FONTS["body"],
                     height=34, width=320).pack(side="left", padx=(0, 10))

        ctk.CTkOptionMenu(fbar, variable=type_var, values=["All", "IN", "OUT"],
                          fg_color=COLORS["card"], button_color=COLORS["border"],
                          text_color=COLORS["txt"], font=FONTS["body"],
                          dropdown_fg_color=COLORS["card"],
                          height=34, width=100,
                          command=lambda _: _filter()).pack(side="left")

        count_lbl = ctk.CTkLabel(fbar, text=f"{len(data)} records",
                                 font=FONTS["small"], text_color=COLORS["txt3"])
        count_lbl.pack(side="right")

        COLS = [
            {"id": "type",    "label": "Type",       "width": 55,
                "stretch": False, "anchor": "center"},
            {"id": "ts",      "label": "Date / Time",
                "width": 145, "stretch": False},
            {"id": "sku",     "label": "SKU",
                "width": 120, "stretch": False},
            {"id": "part",    "label": "Part Name",  "width": 200},
            {"id": "qty",     "label": "Qty",        "width": 55,
                "stretch": False, "anchor": "center"},
            {"id": "unit",    "label": "Unit",       "width": 55,
                "stretch": False, "anchor": "center"},
            {"id": "cost",    "label": "Unit Cost",
                "width": 95,  "stretch": False, "anchor": "e"},
            {"id": "sup",     "label": "Supplier / Reason", "width": 180},
            {"id": "ref",     "label": "Reference",
                "width": 130, "stretch": False},
            {"id": "user",    "label": "User",
                "width": 100, "stretch": False},
        ]
        tbl = DataTable(self._table_frame(), COLS, height=16)
        tbl.pack(fill="both", expand=True)

        all_rows = []
        for i, d in enumerate(data):
            all_rows.append({"id": i, "values": (
                d["type"],
                d["timestamp"][:19].replace(
                    "T", "  ") if d["timestamp"] else "—",
                d["sku"],
                d["part"],
                str(d["quantity"]),
                d["unit"],
                f"₱{d['unit_cost']:,.2f}" if d["unit_cost"] else "—",
                d["supplier"],
                d["reference"],
                d["user"],
            ), "_tag": ""})

        tbl.load(all_rows)

        def _filter(*_):
            term = filter_var.get().strip().lower()
            typ = type_var.get()
            filtered = all_rows
            if typ != "All":
                filtered = [r for r in filtered if r["values"][0] == typ]
            if term:
                filtered = [r for r in filtered if any(
                    term in str(v).lower() for v in r["values"]
                )]
            tbl.load(filtered)
            count_lbl.configure(text=f"{len(filtered)} records")

        debounce = Debouncer(self, delay_ms=160)
        filter_var.trace_add("write", lambda *_: debounce.call(_filter))

    # ── Tab: Audit Log ────────────────────────────────────────────────
    def _render_audit(self):
        db = get_session()
        try:
            logs = AuditService(db).get_all(limit=1000, **self._date_kwargs())
        finally:
            db.close()

        self._stat_cards(self._content, [
            ("Total Log Entries", str(len(logs)), COLORS["txt"]),
            ("Immutable",         "Read-only",    COLORS["green"]),
        ])

        filter_frame = ctk.CTkFrame(self._content, fg_color="transparent")
        filter_frame.pack(fill="x", padx=16, pady=(0, 8))
        filter_var = tk.StringVar()
        ctk.CTkEntry(filter_frame, textvariable=filter_var,
                     placeholder_text="🔍  Filter audit log…",
                     fg_color=COLORS["card"], border_color=COLORS["border"],
                     text_color=COLORS["txt"], font=FONTS["body"],
                     height=34, width=320).pack(side="left")
        count_lbl = ctk.CTkLabel(filter_frame, text=f"{len(logs)} entries",
                                 font=FONTS["small"], text_color=COLORS["txt3"])
        count_lbl.pack(side="right")

        COLS = [
            {"id": "ts",     "label": "Date / Time",
                "width": 145, "stretch": False},
            {"id": "action", "label": "Action",
                "width": 110, "stretch": False},
            {"id": "sku",    "label": "SKU",
                "width": 120, "stretch": False},
            {"id": "part",   "label": "Part Name",    "width": 200},
            {"id": "delta",  "label": "Change",       "width": 70,
                "stretch": False, "anchor": "center"},
            {"id": "user",   "label": "User",
                "width": 100, "stretch": False},
            {"id": "reason", "label": "Reason",       "width": 260},
        ]
        tbl = DataTable(self._table_frame(), COLS, height=16)
        tbl.pack(fill="both", expand=True)

        all_rows = []
        for l in logs:
            delta_str = ""
            if l.delta is not None:
                delta_str = f"+{l.delta}" if l.delta > 0 else str(l.delta)
            all_rows.append({"id": l.id, "values": (
                l.created_at[:19].replace("T", "  ") if l.created_at else "—",
                l.action,
                l.part.sku if l.part else "—",
                l.part.name if l.part else "—",
                delta_str or "—",
                l.user,
                l.reason or "—",
            ), "_tag": ""})

        tbl.load(all_rows)

        def _filter(*_):
            term = filter_var.get().strip().lower()
            filtered = all_rows if not term else [
                r for r in all_rows if any(term in str(v).lower() for v in r["values"])
            ]
            tbl.load(filtered)
            count_lbl.configure(text=f"{len(filtered)} entries")

        debounce = Debouncer(self, delay_ms=160)
        filter_var.trace_add("write", lambda *_: debounce.call(_filter))

    # ── Tab: Adjustments ─────────────────────────────────────────────
    def _render_adjustments(self):
        from core.services.adjustment_service import AdjustmentService, REASONS
        db = get_session()
        try:
            rows = AdjustmentService(db).get_history(limit=1000, **self._date_kwargs())
            data = [{
                "ts": r.created_at, "sku": r.part.sku if r.part else "—",
                "name": r.part.name if r.part else "—",
                "reason": REASONS.get(r.reason_code, r.reason_code),
                "prev": r.previous_count, "new": r.new_count,
                "delta": r.delta, "value": r.value_delta,
                "user": r.user, "note": r.note or "—",
            } for r in rows]
        finally:
            db.close()

        loss = sum(d["value"] for d in data if d["value"] < 0)
        gain = sum(d["value"] for d in data if d["value"] > 0)
        self._stat_cards(self._content, [
            ("Adjustments", str(len(data)), COLORS["txt"]),
            ("Total Loss", f"₱{abs(loss):,.2f}", COLORS["red"] if loss else COLORS["green"]),
            ("Total Gain", f"₱{gain:,.2f}", COLORS["green"]),
        ])

        filter_frame = ctk.CTkFrame(self._content, fg_color="transparent")
        filter_frame.pack(fill="x", padx=16, pady=(0, 8))
        filter_var = tk.StringVar()
        ctk.CTkEntry(filter_frame, textvariable=filter_var,
                     placeholder_text="🔍  Filter adjustments…",
                     fg_color=COLORS["card"], border_color=COLORS["border"],
                     text_color=COLORS["txt"], font=FONTS["body"],
                     height=34, width=320).pack(side="left")
        count_lbl = ctk.CTkLabel(filter_frame, text=f"{len(data)} entries",
                                 font=FONTS["small"], text_color=COLORS["txt3"])
        count_lbl.pack(side="right")

        COLS = [
            {"id": "ts",     "label": "Date / Time", "width": 145, "stretch": False},
            {"id": "sku",    "label": "SKU",         "width": 120, "stretch": False},
            {"id": "name",   "label": "Part Name",   "width": 180},
            {"id": "reason", "label": "Reason",      "width": 130, "stretch": False},
            {"id": "prev",   "label": "Prev",        "width": 60,  "stretch": False, "anchor": "center"},
            {"id": "new",    "label": "New",         "width": 60,  "stretch": False, "anchor": "center"},
            {"id": "delta",  "label": "Δ Qty",       "width": 70,  "stretch": False, "anchor": "center"},
            {"id": "value",  "label": "Δ Value",     "width": 110, "stretch": False, "anchor": "e"},
            {"id": "user",   "label": "User",        "width": 90,  "stretch": False},
            {"id": "note",   "label": "Note",        "width": 180},
        ]
        tbl = DataTable(self._table_frame(), COLS, height=16)
        tbl.pack(fill="both", expand=True)

        all_rows = []
        for i, d in enumerate(data):
            sign = "+" if d["delta"] > 0 else ""
            all_rows.append({"id": i, "_tag": "low" if d["value"] < 0 else "", "values": (
                d["ts"][:19].replace("T", "  ") if d["ts"] else "—",
                d["sku"], d["name"], d["reason"], d["prev"], d["new"],
                f"{sign}{d['delta']}", f"₱{d['value']:,.2f}", d["user"], d["note"],
            )})
        tbl.load(all_rows)

        def _filter(*_):
            term = filter_var.get().strip().lower()
            filtered = all_rows if not term else [
                r for r in all_rows if any(term in str(v).lower() for v in r["values"])]
            tbl.load(filtered)
            count_lbl.configure(text=f"{len(filtered)} entries")

        debounce = Debouncer(self, delay_ms=160)
        filter_var.trace_add("write", lambda *_: debounce.call(_filter))

    # ── Tab: Returns ─────────────────────────────────────────────────
    def _render_returns(self):
        from core.services.return_service import ReturnService, REASONS, CONDITIONS
        db = get_session()
        try:
            rows = ReturnService(db).get_history(limit=1000, **self._date_kwargs())
            data = [{
                "ts": r.created_at, "sku": r.part.sku if r.part else "—",
                "name": r.part.name if r.part else "—", "qty": r.quantity,
                "cond": CONDITIONS.get(r.condition, r.condition),
                "reason": REASONS.get(r.reason_code, r.reason_code),
                "refund": r.refund_amount, "method": r.refund_method,
                "restock": r.restock_qty, "profit": r.profit_delta, "user": r.user,
            } for r in rows]
        finally:
            db.close()

        refunds = sum(d["refund"] for d in data)
        restocked = sum(d["restock"] for d in data)
        scrapped = sum(d["qty"] - d["restock"] for d in data)
        profit = sum(d["profit"] for d in data)
        self._stat_cards(self._content, [
            ("Returns", str(len(data)), COLORS["txt"]),
            ("Total Refunds", f"₱{refunds:,.2f}", COLORS["amber"]),
            ("Units Restocked", str(restocked), COLORS["green"]),
            ("Units Scrapped", str(scrapped), COLORS["red"] if scrapped else COLORS["green"]),
            ("Profit Impact", f"₱{profit:,.2f}", COLORS["red"] if profit < 0 else COLORS["green"]),
        ])

        filter_frame = ctk.CTkFrame(self._content, fg_color="transparent")
        filter_frame.pack(fill="x", padx=16, pady=(0, 8))
        filter_var = tk.StringVar()
        ctk.CTkEntry(filter_frame, textvariable=filter_var,
                     placeholder_text="🔍  Filter returns…",
                     fg_color=COLORS["card"], border_color=COLORS["border"],
                     text_color=COLORS["txt"], font=FONTS["body"],
                     height=34, width=320).pack(side="left")
        count_lbl = ctk.CTkLabel(filter_frame, text=f"{len(data)} entries",
                                 font=FONTS["small"], text_color=COLORS["txt3"])
        count_lbl.pack(side="right")

        COLS = [
            {"id": "ts",      "label": "Date / Time", "width": 140, "stretch": False},
            {"id": "sku",     "label": "SKU",         "width": 115, "stretch": False},
            {"id": "name",    "label": "Part",        "width": 170},
            {"id": "qty",     "label": "Qty",         "width": 50,  "stretch": False, "anchor": "center"},
            {"id": "cond",    "label": "Condition",   "width": 95,  "stretch": False},
            {"id": "reason",  "label": "Reason",      "width": 120, "stretch": False},
            {"id": "refund",  "label": "Refund",      "width": 95,  "stretch": False, "anchor": "e"},
            {"id": "method",  "label": "Method",      "width": 85,  "stretch": False},
            {"id": "profit",  "label": "Profit Impact", "width": 110, "stretch": False, "anchor": "e"},
            {"id": "user",    "label": "By",          "width": 85,  "stretch": False},
        ]
        tbl = DataTable(self._table_frame(), COLS, height=16)
        tbl.pack(fill="both", expand=True)

        all_rows = []
        for i, d in enumerate(data):
            all_rows.append({"id": i, "_tag": "low" if d["profit"] < 0 else "", "values": (
                (d["ts"] or "")[:19].replace("T", "  "), d["sku"], d["name"], d["qty"],
                d["cond"], d["reason"], f"₱{d['refund']:,.2f}", d["method"],
                f"₱{d['profit']:,.2f}", d["user"])})
        tbl.load(all_rows)

        def _filter(*_):
            term = filter_var.get().strip().lower()
            filtered = all_rows if not term else [
                r for r in all_rows if any(term in str(v).lower() for v in r["values"])]
            tbl.load(filtered)
            count_lbl.configure(text=f"{len(filtered)} entries")

        debounce = Debouncer(self, delay_ms=160)
        filter_var.trace_add("write", lambda *_: debounce.call(_filter))

    # ── Tab: Aging Analysis ───────────────────────────────────────────
    def _render_aging(self):
        db = get_session()
        try:
            data = ReportService(db).aging_analysis(days=90)
        finally:
            db.close()

        total_val = sum(d["value"] for d in data)
        self._stat_cards(self._content, [
            ("Idle Parts (90+ days)",
             str(len(data)),           COLORS["amber"]),
            ("Idle Inventory Value",
             f"₱{total_val:,.2f}",     COLORS["amber"]),
        ])

        ctk.CTkLabel(self._content,
                     text="  ⚠  Parts below have had no stock-out activity in the last 90 days.",
                     font=FONTS["small"], text_color=COLORS["amber"],
                     anchor="w").pack(fill="x", padx=16, pady=(0, 6))

        COLS = [
            {"id": "sku",     "label": "SKU",
                "width": 120, "stretch": False},
            {"id": "name",    "label": "Part Name",    "width": 200},
            {"id": "cat",     "label": "Category",
                "width": 110, "stretch": False},
            {"id": "stock",   "label": "Stock",        "width": 70,
                "stretch": False, "anchor": "center"},
            {"id": "unit",    "label": "Unit",         "width": 55,
                "stretch": False, "anchor": "center"},
            {"id": "cost",    "label": "Unit Cost",
                "width": 95,  "stretch": False, "anchor": "e"},
            {"id": "value",   "label": "Total Value",
                "width": 110, "stretch": False, "anchor": "e"},
            {"id": "bin",     "label": "Bin",
                "width": 80,  "stretch": False},
            {"id": "last",    "label": "Last Issued",
                "width": 110, "stretch": False},
            {"id": "in",      "label": "Total In",     "width": 75,
                "stretch": False, "anchor": "center"},
            {"id": "out",     "label": "Total Out",    "width": 75,
                "stretch": False, "anchor": "center"},
        ]
        tbl = DataTable(self._table_frame(), COLS, height=18)
        tbl.pack(fill="both", expand=True)

        rows = [{"id": i, "values": (
            d["sku"], d["name"], d["category"],
            d["stock"], d["unit"],
            f"₱{d['unit_cost']:,.2f}",
            f"₱{d['value']:,.2f}",
            d["bin"], d["last_issued"],
            d["total_in"], d["total_out"],
        ), "_tag": "warn"} for i, d in enumerate(data)]
        tbl.load(rows)

    # ── Export ────────────────────────────────────────────────────────
    def _export(self):
        if self._active_tab == "returns":
            try:
                path = self._export_returns_excel()
            except Exception as e:
                Toast(self.app, f"Export failed: {e}", kind="error")
                return
            Toast(self.app, f"Exported Excel: {os.path.basename(path)}", kind="success")
            self._open_path(path)
            return

        if self._active_tab == "adjustments":
            try:
                path = self._export_adjustments_excel()
            except Exception as e:
                Toast(self.app, f"Export failed: {e}", kind="error")
                return
            Toast(self.app, f"Exported Excel: {os.path.basename(path)}", kind="success")
            self._open_path(path)
            return

        export_map = {
            "inventory": "inventory",
            "category":  "category",
            "movements": "movements",
            "audit":     "audit",
            "aging":     "aging",
            "sales":     "sales",
        }
        export_type = export_map.get(self._active_tab, "inventory")

        # Parts Inventory and By Category are current snapshots — never date-filtered.
        date_kwargs = ({} if self._active_tab in ("inventory", "category")
                       else self._date_kwargs())
        # Honour the "Low Stock only" filter when exporting Parts Inventory.
        if self._active_tab == "inventory" and self._inv_low_only:
            date_kwargs["low_only"] = True

        db = get_session()
        try:
            svc = ReportService(db)

            # Main path: export with the selected Report Date.
            # Fallback path: if an older report_service.py is still loaded,
            # export without date arguments instead of breaking all tabs.
            try:
                path = svc.export_to_excel(export_type, **date_kwargs)
            except TypeError:
                path = svc.export_to_excel(export_type)

        except Exception as e:
            Toast(self.app, f"Export failed: {e}", kind="error")
            return
        finally:
            db.close()

        Toast(
            self.app, f"Exported Excel: {os.path.basename(path)}", kind="success")
        self._open_path(path)

    def _export_audit_excel(self):
        from config.settings import EXPORT_DIR
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        os.makedirs(EXPORT_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(EXPORT_DIR, f"inventra_audit_{ts}.xlsx")

        db = get_session()
        try:
            logs = AuditService(db).get_all(limit=10000, **self._date_kwargs())
        finally:
            db.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Log"

        headers = ["Date/Time", "Action", "SKU",
                   "Part Name", "Change", "User", "Reason"]

        hdr_fill = PatternFill("solid", fgColor="1D3461")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center")
        alt_fill = PatternFill("solid", fgColor="F7F6F3")
        border = Border(bottom=Side(style="thin", color="E5E5E5"))

        ws.row_dimensions[1].height = 32
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = Border(
                bottom=Side(style="medium", color="FFFFFF"),
                right=Side(style="thin", color="2A4A82"),
            )

        for ri, log in enumerate(logs, 2):
            delta = ""
            if log.delta is not None:
                delta = f"+{log.delta}" if log.delta > 0 else str(log.delta)

            values = [
                log.created_at[:19].replace(
                    "T", " ") if log.created_at else "",
                log.action,
                log.part.sku if log.part else "",
                log.part.name if log.part else "",
                delta,
                log.user,
                log.reason or "",
            ]

            fill = alt_fill if ri % 2 == 0 else PatternFill(
                "solid", fgColor="FFFFFF")
            for ci, value in enumerate(values, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.fill = fill
                cell.border = border

        for ci, header in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            max_len = len(str(header))
            for ri in range(2, len(logs) + 2):
                value = ws.cell(row=ri, column=ci).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 42)

        ws.freeze_panes = "A2"
        wb.save(filename)
        return filename

    def _export_adjustments_excel(self):
        from config.settings import EXPORT_DIR
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from core.services.adjustment_service import AdjustmentService, REASONS

        os.makedirs(EXPORT_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(EXPORT_DIR, f"inventra_adjustments_{ts}.xlsx")

        db = get_session()
        try:
            rows = AdjustmentService(db).get_history(limit=10000, **self._date_kwargs())
            data = [(
                r.created_at[:19].replace("T", " ") if r.created_at else "",
                r.part.sku if r.part else "", r.part.name if r.part else "",
                REASONS.get(r.reason_code, r.reason_code),
                r.previous_count, r.new_count, r.delta, r.value_delta,
                r.user, r.note or "",
            ) for r in rows]
        finally:
            db.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Adjustments"
        headers = ["Date/Time", "SKU", "Part Name", "Reason", "Prev", "New",
                   "Delta Qty", "Delta Value", "User", "Note"]
        hdr_fill = PatternFill("solid", fgColor="1D3461")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        ws.row_dimensions[1].height = 30
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        border = Border(bottom=Side(style="thin", color="E5E5E5"))
        for ri, values in enumerate(data, 2):
            for ci, value in enumerate(values, 1):
                c = ws.cell(row=ri, column=ci, value=value)
                c.border = border
        for ci, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(
                12, min(len(header) + 4, 40))
        ws.freeze_panes = "A2"
        wb.save(filename)
        return filename

    def _export_returns_excel(self):
        from config.settings import EXPORT_DIR
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from core.services.return_service import ReturnService, REASONS, CONDITIONS

        os.makedirs(EXPORT_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(EXPORT_DIR, f"inventra_returns_{ts}.xlsx")

        db = get_session()
        try:
            rows = ReturnService(db).get_history(limit=10000, **self._date_kwargs())
            data = [(
                (r.created_at or "")[:19].replace("T", " "),
                r.part.sku if r.part else "", r.part.name if r.part else "",
                r.quantity, CONDITIONS.get(r.condition, r.condition),
                REASONS.get(r.reason_code, r.reason_code), r.refund_amount,
                r.refund_method, r.restock_qty, r.profit_delta, r.user,
            ) for r in rows]
        finally:
            db.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Returns"
        headers = ["Date/Time", "SKU", "Part", "Qty", "Condition", "Reason",
                   "Refund", "Method", "Restocked", "Profit Impact", "By"]
        hdr_fill = PatternFill("solid", fgColor="1D3461")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        ws.row_dimensions[1].height = 30
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        border = Border(bottom=Side(style="thin", color="E5E5E5"))
        for ri, values in enumerate(data, 2):
            for ci, value in enumerate(values, 1):
                c = ws.cell(row=ri, column=ci, value=value)
                c.border = border
        for ci, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(
                12, min(len(header) + 4, 40))
        ws.freeze_panes = "A2"
        wb.save(filename)
        return filename

    def _open_path(self, path: str):
        folder = os.path.dirname(path)
        try:
            if sys.platform == "win32":
                subprocess.Popen(f'explorer /select,"{path}"')
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception:
            pass

    # ── Tab: Sales ────────────────────────────────────────────────────
    def _render_sales(self):
        from core.services.report_service import ReportService as RS
        date_kwargs = self._date_kwargs()
        db = get_session()
        try:
            stats = RS(db).sales_summary_stats(**date_kwargs)
            detail = RS(db).sales_detail(**date_kwargs)
            by_part = RS(db).sales_by_part(**date_kwargs)
        finally:
            db.close()

        # ── Summary cards ─────────────────────────────────────────────
        self._stat_cards(self._content, [
            ("Transactions",   str(
                stats["transactions"]),            COLORS["txt"]),
            ("Units Sold",     str(
                stats["units_sold"]),              COLORS["blue"]),
            ("Net Revenue",
             f"₱{stats['net_revenue']:,.2f}",       COLORS["navy"]),
            ("Discounts Given",
             f"₱{stats['total_discounts']:,.2f}",   COLORS["amber"]),
            ("Refunds", f"₱{stats.get('refunds', 0):,.2f}",
             COLORS["red"] if stats.get('refunds') else COLORS["green"]),
            ("Fees", f"₱{stats.get('labor', 0):,.2f}",
             COLORS["navy"] if stats.get('labor') else COLORS["txt3"]),
            ("Gross Profit",   f"₱{stats['gross_profit']:,.2f}",
             COLORS["green"] if stats["gross_profit"] >= 0 else COLORS["red"]),
            ("Margin",         f"{stats['margin_pct']:.1f}%",
             COLORS["green"] if stats["margin_pct"] >= 0 else COLORS["red"]),
        ])

        # ── Sub-tab: Transactions / By Part ───────────────────────────
        sub_var = tk.StringVar(value="transactions")
        sub_bar = ctk.CTkFrame(self._content, fg_color="transparent")
        sub_bar.pack(fill="x", padx=16, pady=(0, 6))

        sub_btns = {}
        sub_content = ctk.CTkFrame(
            self._content, fg_color=COLORS["card"], corner_radius=0)
        sub_content.pack(fill="both", expand=True)

        def _switch_sub(key):
            sub_var.set(key)
            for k, b in sub_btns.items():
                b.configure(
                    fg_color=COLORS["navy"] if k == key else COLORS["bg2"],
                    text_color="#FFFFFF" if k == key else COLORS["txt2"],
                    hover_color=COLORS["navy_hover"] if k == key else COLORS["border"],
                )
            for w in sub_content.winfo_children():
                w.destroy()
            if key == "transactions":
                _render_transactions(sub_content)
            else:
                _render_by_part(sub_content)

        for key, label in [("transactions", "All Transactions"), ("by_part", "By Part")]:
            b = ctk.CTkButton(sub_bar, text=label,
                              fg_color=COLORS["navy"] if key == "transactions" else COLORS["bg2"],
                              hover_color=COLORS["navy_hover"] if key == "transactions" else COLORS["border"],
                              text_color="#FFFFFF" if key == "transactions" else COLORS["txt2"],
                              font=FONTS["body"], height=30, width=140, corner_radius=8,
                              command=lambda k=key: _switch_sub(k))
            b.pack(side="left", padx=(0, 6))
            sub_btns[key] = b

        # ── Transactions table ────────────────────────────────────────
        def _render_transactions(parent):
            fbar = ctk.CTkFrame(parent, fg_color=COLORS["bg2"])
            fbar.pack(fill="x", padx=0, pady=0)
            inner = ctk.CTkFrame(fbar, fg_color="transparent")
            inner.pack(fill="x", padx=12, pady=8)

            filter_var = tk.StringVar()
            ctk.CTkEntry(inner, textvariable=filter_var,
                         placeholder_text="🔍  Filter by part, SKU, reason…",
                         fg_color=COLORS["card"], border_color=COLORS["border"],
                         text_color=COLORS["txt"], font=FONTS["body"],
                         height=32, width=300).pack(side="left", padx=(0, 10))

            count_lbl = ctk.CTkLabel(inner, text=f"{len(detail)} transactions",
                                     font=FONTS["small"], text_color=COLORS["txt3"])
            count_lbl.pack(side="right")

            COLS = [
                {"id": "date",    "label": "Date / Time",
                    "width": 145, "stretch": False},
                {"id": "sku",     "label": "SKU",
                    "width": 115, "stretch": False},
                {"id": "name",    "label": "Part Name",     "width": 180},
                {"id": "cat",     "label": "Category",
                    "width": 100, "stretch": False},
                {"id": "qty",     "label": "Qty",           "width": 50,
                    "stretch": False, "anchor": "center"},
                {"id": "price",   "label": "Unit Price",
                    "width": 95,  "stretch": False, "anchor": "e"},
                {"id": "disc",    "label": "Disc %",        "width": 60,
                    "stretch": False, "anchor": "center"},
                {"id": "disc_a",  "label": "Disc Amt",
                    "width": 90,  "stretch": False, "anchor": "e"},
                {"id": "subtot",  "label": "Subtotal",
                    "width": 95,  "stretch": False, "anchor": "e"},
                {"id": "total",   "label": "Total",
                    "width": 100, "stretch": False, "anchor": "e"},
                {"id": "cost",    "label": "Unit Cost",
                    "width": 90,  "stretch": False, "anchor": "e"},
                {"id": "profit",  "label": "Profit",
                    "width": 95,  "stretch": False, "anchor": "e"},
                {"id": "margin",  "label": "Margin %",      "width": 80,
                    "stretch": False, "anchor": "center"},
                {"id": "reason",  "label": "Reason",
                    "width": 120, "stretch": False},
                {"id": "jobref",  "label": "Job Ref",
                    "width": 100, "stretch": False},
                {"id": "user",    "label": "User",
                    "width": 90,  "stretch": False},
            ]
            tbl = DataTable(parent, COLS, height=16)
            tbl.pack(fill="both", expand=True)

            all_rows = []
            for i, d in enumerate(detail):
                tag = "low" if d["gross_profit"] < 0 else ""
                all_rows.append({"id": i, "_tag": tag, "values": (
                    d["issued_at"],
                    d["sku"],
                    d["name"],
                    d["category"],
                    d["quantity"],
                    f"₱{d['selling_price']:,.2f}",
                    f"{d['discount_pct']:.0f}%" if d["discount_pct"] else "—",
                    f"₱{d['discount_amount']:,.2f}" if d["discount_amount"] else "—",
                    f"₱{d['subtotal']:,.2f}",
                    f"₱{d['total_amount']:,.2f}",
                    f"₱{d['unit_cost']:,.2f}",
                    f"₱{d['gross_profit']:,.2f}",
                    f"{d['margin_pct']:.1f}%",
                    d["reason"],
                    d["job_ref"],
                    d["issued_by"],
                )})

            tbl.load(all_rows)

            def _filter(*_):
                term = filter_var.get().strip().lower()
                filtered = all_rows if not term else [
                    r for r in all_rows if any(term in str(v).lower() for v in r["values"])
                ]
                tbl.load(filtered)
                count_lbl.configure(text=f"{len(filtered)} transactions")

            debounce = Debouncer(self, delay_ms=160)
            filter_var.trace_add("write", lambda *_: debounce.call(_filter))

        # ── By Part summary table ─────────────────────────────────────
        def _render_by_part(parent):
            COLS = [
                {"id": "sku",     "label": "SKU",
                    "width": 120, "stretch": False},
                {"id": "name",    "label": "Part Name",   "width": 200},
                {"id": "cat",     "label": "Category",
                    "width": 110, "stretch": False},
                {"id": "qty",     "label": "Qty Sold",    "width": 80,
                    "stretch": False, "anchor": "center"},
                {"id": "unit",    "label": "Unit",        "width": 55,
                    "stretch": False, "anchor": "center"},
                {"id": "avg",     "label": "Avg Price",
                    "width": 100, "stretch": False, "anchor": "e"},
                {"id": "rev",     "label": "Revenue",
                    "width": 120, "stretch": False, "anchor": "e"},
                {"id": "profit",  "label": "Profit",
                    "width": 110, "stretch": False, "anchor": "e"},
                {"id": "margin",  "label": "Margin %",    "width": 85,
                    "stretch": False, "anchor": "center"},
                {"id": "txns",    "label": "# Txns",      "width": 65,
                    "stretch": False, "anchor": "center"},
            ]
            tbl = DataTable(parent, COLS, height=18)
            tbl.pack(fill="both", expand=True)

            total_rev = sum(d["revenue"] for d in by_part)
            total_profit = sum(d["profit"] for d in by_part)
            rows = []
            for i, d in enumerate(by_part):
                tag = "low" if d["profit"] < 0 else ""
                rows.append({"id": i, "_tag": tag, "values": (
                    d["sku"], d["name"], d["category"],
                    d["qty_sold"], d["unit"],
                    f"₱{d['avg_price']:,.2f}",
                    f"₱{d['revenue']:,.2f}",
                    f"₱{d['profit']:,.2f}",
                    f"{d['margin_pct']:.1f}%",
                    d["txn_count"],
                )})

            # Totals row
            rows.append({"id": -1, "_tag": "", "values": (
                "", "TOTAL", "",
                sum(d["qty_sold"] for d in by_part), "",
                "",
                f"₱{total_rev:,.2f}",
                f"₱{total_profit:,.2f}",
                f"{(total_profit/max(total_rev, 1)*100):.1f}%",
                sum(d["txn_count"] for d in by_part),
            )})
            tbl.load(rows)

        _switch_sub("transactions")

    # ── Sales Excel export compatibility ──────────────────────────────
    def _export_sales_csv(self):
        """Backward-compatible method name. Sales now exports XLSX only."""
        db = get_session()
        try:
            svc = ReportService(db)
            try:
                path = svc.export_to_excel("sales", **self._date_kwargs())
            except TypeError:
                path = svc.export_to_excel("sales")
        except Exception as e:
            Toast(self.app, f"Export failed: {e}", kind="error")
            return
        finally:
            db.close()

        Toast(
            self.app, f"Exported Excel: {os.path.basename(path)}", kind="success")
        self._open_path(path)
