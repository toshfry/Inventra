"""
Inventra Web Server
Run with: python web_server.py
Access on Android/other devices: http://<your-local-ip>:5000
"""

import socket
from database.models.category import Category
from core.validators.supplier_schema import SupplierCreate, SupplierUpdate
from core.validators.transaction_schema import StockInCreate, StockOutCreate
from core.validators.part_schema import PartCreate, PartUpdate
from core.services.audit_service import AuditService
from core.services.dashboard_service import DashboardService
from core.services.supplier_service import SupplierService
from core.services.stock_service import StockService
from core.services.parts_service import PartsService
from core.services.pos_service import PosService
from core.services.settings_service import SettingsService
from core.services.receipt_renderer import render_receipt_html
from core.validators.pos_schema import SaleCreate, PosSettingsUpdate
from core.services.auth_service import AuthService, set_current_user, logout
from database.engine import init_db, get_session
from core.licensing.license_manager import activate_from_key, get_computer_id, license_status, validate_license
from functools import wraps
from flask import Flask, request, jsonify, session, send_from_directory
import sys
import os

# Support running as a PyInstaller frozen exe
if getattr(sys, "frozen", False):
    _BASE = sys._MEIPASS
else:
    _BASE = os.path.dirname(os.path.abspath(__file__))

sys.path.insert(0, _BASE)


_WEB_DIR = os.path.join(_BASE, "web")
app = Flask(__name__, static_folder=os.path.join(_WEB_DIR, "static"),
            template_folder=os.path.join(_WEB_DIR, "templates"))
app.secret_key = os.environ.get(
    "INVENTRA_WEB_SECRET_KEY",
    "inventra-local-dev-secret-change-me",
)

# ── Reports / Exports directory ───────────────────────────────────────────────
try:
    from config.settings import EXPORT_DIR as _CONFIG_EXPORT_DIR
except Exception:
    _CONFIG_EXPORT_DIR = None

_REPORTS_DIR = os.path.abspath(
    _CONFIG_EXPORT_DIR or os.path.join(_BASE, "exports"))
REPORT_EXTENSIONS = {".xlsx", ".xls", ".csv", ".pdf"}


# ── Helpers ──────────────────────────────────────────────────────────────────


def ok(data=None, **kwargs):
    resp = {"ok": True}
    if data is not None:
        resp["data"] = data
    resp.update(kwargs)
    return jsonify(resp)


def err(msg, code=400):
    return jsonify({"ok": False, "error": str(msg)}), code


# ── Activation guard ──────────────────────────────────────────────────────────
@app.before_request
def require_activation_for_api():
    """
    Protect web/API mode too. The desktop app shows the full activation window.
    These API endpoints are here so the web server cannot be used normally
    without a valid local license file.
    """
    path = request.path or ""

    if not path.startswith("/api/"):
        return None

    if path.startswith("/api/activation/"):
        return None

    valid, message, _ = validate_license()
    if not valid:
        return err(f"{message} Computer ID: {get_computer_id()}", 403)

    return None


@app.route("/api/activation/status", methods=["GET"])
def api_activation_status():
    return ok(license_status())


@app.route("/api/activation/activate", methods=["POST"])
def api_activation_activate():
    data = request.get_json() or {}
    key = data.get("activation_key") or data.get("key") or ""
    valid, message, payload = activate_from_key(key)

    if not valid:
        return err(message, 400)

    return ok({
        "message": message,
        "license": payload,
        "status": license_status(),
    })


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return err("Unauthorized", 401)
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return err("Unauthorized", 401)
        if not session.get("is_admin"):
            return err("Admin access required", 403)
        return f(*args, **kwargs)
    return decorated


def current_user():
    return session.get("username", "system")


def current_display_name():
    """Human-friendly name for receipts (full name, falling back to username)."""
    return session.get("full_name") or session.get("username", "system")

# ── Static / SPA ─────────────────────────────────────────────────────────────


@app.route("/logo.png")
def logo():
    return send_from_directory(_WEB_DIR, "logo.png")


@app.route("/")
@app.route("/<path:path>")
def index(path=""):
    # Serve the single-page app for all non-API routes
    if path.startswith("api/") or path.startswith("static/"):
        return err("Not found", 404)
    return send_from_directory(_WEB_DIR, "index.html")

# ── Auth ──────────────────────────────────────────────────────────────────────


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return err("Username and password required")
    db = get_session()
    try:
        svc = AuthService(db)
        user = svc.login(username, password)
        session["user_id"] = user.id
        session["username"] = user.username
        session["full_name"] = user.full_name
        session["is_admin"] = user.is_admin
        set_current_user(user)
        return ok({"username": user.username, "full_name": user.full_name,
                   "role": user.role, "is_admin": user.is_admin})
    except ValueError as e:
        return err(str(e), 401)
    finally:
        db.close()


@app.route("/api/auth/logout", methods=["POST"])
def do_logout():
    session.clear()
    logout()
    return ok()


@app.route("/api/auth/me")
def me():
    if not session.get("user_id"):
        return err("Not logged in", 401)
    return ok({"username":  session["username"],
               "full_name": session["full_name"],
               "is_admin":  session["is_admin"]})

# ── Categories ────────────────────────────────────────────────────────────────


@app.route("/api/categories", methods=["GET"])
@login_required
def list_categories():
    db = get_session()
    try:
        cats = db.query(Category).order_by(Category.name).all()
        return ok([{"id": c.id, "name": c.name, "color_hex": c.color_hex} for c in cats])
    finally:
        db.close()


@app.route("/api/categories", methods=["POST"])
@admin_required
def create_category():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    color = (data.get("color_hex") or "#888888").strip()
    if not name:
        return err("Category name required")
    db = get_session()
    try:
        existing = db.query(Category).filter(Category.name == name).first()
        if existing:
            return err(f"Category '{name}' already exists")
        cat = Category(name=name, color_hex=color)
        db.add(cat)
        db.commit()
        db.refresh(cat)
        return ok({"id": cat.id, "name": cat.name, "color_hex": cat.color_hex})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/categories/<int:cat_id>", methods=["PUT"])
@admin_required
def update_category(cat_id):
    data = request.get_json() or {}
    db = get_session()
    try:
        cat = db.get(Category, cat_id)
        if not cat:
            return err("Category not found", 404)
        if "name" in data and data["name"].strip():
            # Check uniqueness
            existing = db.query(Category).filter(
                Category.name == data["name"].strip(),
                Category.id != cat_id
            ).first()
            if existing:
                return err(f"Category '{data['name']}' already exists")
            cat.name = data["name"].strip()
        if "color_hex" in data:
            cat.color_hex = data["color_hex"].strip() or "#888888"
        db.commit()
        db.refresh(cat)
        return ok({"id": cat.id, "name": cat.name, "color_hex": cat.color_hex})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/categories/<int:cat_id>", methods=["DELETE"])
@admin_required
def delete_category(cat_id):
    db = get_session()
    try:
        cat = db.get(Category, cat_id)
        if not cat:
            return err("Category not found", 404)
        if cat.parts:
            return err("Cannot delete a category that has parts assigned to it. Re-assign the parts first.")
        db.delete(cat)
        db.commit()
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Parts ─────────────────────────────────────────────────────────────────────


@app.route("/api/parts", methods=["GET"])
@login_required
def list_parts():
    search = request.args.get("search", "")
    category = request.args.get("category", "")
    db = get_session()
    try:
        svc = PartsService(db)
        rows = svc.get_stock_view(search=search, category=category)
        return ok(rows)
    finally:
        db.close()


@app.route("/api/parts/<int:part_id>", methods=["GET"])
@login_required
def get_part(part_id):
    db = get_session()
    try:
        svc = PartsService(db)
        part = svc.get_by_id(part_id)
        if not part:
            return err("Part not found", 404)
        return ok({
            "id": part.id, "sku": part.sku, "name": part.name,
            "description": part.description, "category_id": part.category_id,
            "category": part.category.name if part.category else None,
            "unit_cost": part.unit_cost, "selling_price": part.selling_price,
            "unit": part.unit, "min_stock": part.min_stock,
            "bin_location": part.bin_location, "vehicle_makes": part.vehicle_makes,
            "current_stock": part.current_stock, "is_active": part.is_active,
        })
    finally:
        db.close()


@app.route("/api/parts", methods=["POST"])
@login_required
def create_part():
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = PartCreate(**data)
        svc = PartsService(db)
        part = svc.create(schema, user=current_user())
        return ok({"id": part.id, "sku": part.sku, "name": part.name}), 201
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/parts/<int:part_id>", methods=["PUT"])
@login_required
def update_part(part_id):
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = PartUpdate(**data)
        svc = PartsService(db)
        part = svc.update(part_id, schema, user=current_user())
        return ok({"id": part.id, "sku": part.sku, "name": part.name})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/parts/<int:part_id>", methods=["DELETE"])
@admin_required
def delete_part(part_id):
    """
    Delete from the web Parts Library.

    This uses soft-delete/deactivate so parts with stock history are still
    removed from the active Parts Library without breaking transaction history.
    """
    db = get_session()
    try:
        PartsService(db).deactivate(part_id, user=current_user())
        DashboardService(db).invalidate()
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/parts/<int:part_id>/deactivate", methods=["POST"])
@login_required
def deactivate_part(part_id):
    db = get_session()
    try:
        PartsService(db).deactivate(part_id, user=current_user())
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/parts/<int:part_id>/reactivate", methods=["POST"])
@login_required
def reactivate_part(part_id):
    db = get_session()
    try:
        PartsService(db).reactivate(part_id, user=current_user())
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Stock Adjustments ──────────────────────────────────────────────────────────


@app.route("/api/parts/<int:part_id>/adjust", methods=["POST"])
@admin_required
def adjust_part_stock(part_id):
    from core.services.adjustment_service import AdjustmentService
    from core.validators.adjustment_schema import AdjustmentCreate
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = AdjustmentCreate(
            part_id=part_id,
            mode=data.get("mode", "set"),
            value=int(data.get("value", 0)),
            reason_code=data.get("reason_code", "OTHER"),
            note=(data.get("note") or None),
        )
        adj = AdjustmentService(db).adjust(schema, user=current_user())
        DashboardService(db).invalidate()
        return ok({"id": adj.id, "new_count": adj.new_count, "delta": adj.delta})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/adjustments", methods=["GET"])
@login_required
def list_adjustments():
    from core.services.adjustment_service import AdjustmentService, REASONS
    date_from = request.args.get("date_from") or None
    date_to = request.args.get("date_to") or None
    db = get_session()
    try:
        rows = AdjustmentService(db).get_history(
            date_from=date_from, date_to=date_to, limit=1000)
        return ok([{
            "id": r.id,
            "sku": r.part.sku if r.part else "—",
            "name": r.part.name if r.part else "—",
            "delta": r.delta, "previous_count": r.previous_count,
            "new_count": r.new_count, "reason_code": r.reason_code,
            "reason": REASONS.get(r.reason_code, r.reason_code),
            "note": r.note, "value_delta": r.value_delta,
            "user": r.user, "created_at": r.created_at,
        } for r in rows])
    finally:
        db.close()


# ── Customer Returns ────────────────────────────────────────────────────────────


@app.route("/api/returns", methods=["GET"])
@login_required
def list_returns():
    from core.services.return_service import ReturnService, REASONS, CONDITIONS
    date_from = request.args.get("date_from") or None
    date_to = request.args.get("date_to") or None
    db = get_session()
    try:
        rows = ReturnService(db).get_history(
            date_from=date_from, date_to=date_to, limit=1000)
        return ok([{
            "id": r.id,
            "sku": r.part.sku if r.part else "—",
            "name": r.part.name if r.part else "—",
            "quantity": r.quantity, "condition": r.condition,
            "condition_label": CONDITIONS.get(r.condition, r.condition),
            "reason_code": r.reason_code,
            "reason": REASONS.get(r.reason_code, r.reason_code),
            "refund_amount": r.refund_amount, "refund_method": r.refund_method,
            "restock_qty": r.restock_qty, "profit_delta": r.profit_delta,
            "user": r.user, "created_at": r.created_at,
        } for r in rows])
    finally:
        db.close()


@app.route("/api/returns/issues", methods=["GET"])
@login_required
def list_returnable_issues():
    from core.services.return_service import ReturnService
    search = request.args.get("search", "")
    db = get_session()
    try:
        return ok(ReturnService(db).get_returnable_issues(search=search, limit=100))
    finally:
        db.close()


@app.route("/api/returns", methods=["POST"])
@login_required
def create_return():
    from core.services.return_service import ReturnService
    from core.validators.return_schema import ReturnCreate
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = ReturnCreate(**data)
        ret = ReturnService(db).process_return(schema, user=current_user())
        DashboardService(db).invalidate()
        return ok({"id": ret.id}), 201
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


# ── Stock In ──────────────────────────────────────────────────────────────────


@app.route("/api/stock/in", methods=["GET"])
@login_required
def list_stock_in():
    part_id = request.args.get("part_id", type=int)
    date_from = request.args.get("date_from") or None
    date_to = request.args.get("date_to") or None
    db = get_session()
    try:
        svc = StockService(db)
        rows = svc.get_stock_in_history(part_id=part_id, limit=200,
                                        date_from=date_from, date_to=date_to)
        result = []
        for r in rows:
            result.append({
                "id": r.id, "part_id": r.part_id,
                "part_name": r.part.name if r.part else "—",
                "supplier_id": r.supplier_id,
                "supplier_name": r.supplier.name if r.supplier else "—",
                "quantity": r.quantity, "unit_cost": r.unit_cost,
                "reference_no": r.reference_no, "notes": r.notes,
                "received_by": r.received_by, "received_at": r.received_at,
            })
        return ok(result)
    finally:
        db.close()


@app.route("/api/stock/in", methods=["POST"])
@login_required
def receive_stock():
    data = request.get_json() or {}
    data.setdefault("received_by", current_user())
    db = get_session()
    try:
        schema = StockInCreate(**data)
        txn = StockService(db).receive_stock(schema)
        DashboardService(db).invalidate()
        return ok({"id": txn.id, "quantity": txn.quantity}), 201
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/stock/in/<int:txn_id>", methods=["DELETE"])
@admin_required
def cancel_stock_in(txn_id):
    db = get_session()
    try:
        StockService(db).cancel_stock_in(txn_id, user=current_user())
        DashboardService(db).invalidate()
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Stock Out ─────────────────────────────────────────────────────────────────


@app.route("/api/stock/out", methods=["GET"])
@login_required
def list_stock_out():
    part_id = request.args.get("part_id", type=int)
    date_from = request.args.get("date_from") or None
    date_to = request.args.get("date_to") or None
    db = get_session()
    try:
        svc = StockService(db)
        rows = svc.get_stock_out_history(part_id=part_id, limit=200,
                                         date_from=date_from, date_to=date_to)

        # Units of each sale that have been returned (for the Returned column).
        from database.models.customer_return import CustomerReturn
        from sqlalchemy import func
        ret_map = dict(
            db.query(CustomerReturn.stock_out_id,
                     func.sum(CustomerReturn.quantity))
            .filter(CustomerReturn.stock_out_id.isnot(None))
            .group_by(CustomerReturn.stock_out_id).all())

        # Tax factor per POS-linked stock-out (grand_total / net) so net_total
        # can be shown tax-inclusive (matches the receipt); profit stays ex-tax.
        from database.models.sale_item import SaleItem
        from database.models.sale import Sale
        tax_factor = {}
        for so_id, gt, sub, disc in (
                db.query(SaleItem.stock_out_id, Sale.grand_total, Sale.subtotal,
                         Sale.discount_total)
                .join(Sale, SaleItem.sale_id == Sale.id)
                .filter(SaleItem.stock_out_id.isnot(None)).all()):
            net = (sub or 0) - (disc or 0)
            if so_id and net > 0 and gt:
                tax_factor[so_id] = gt / net

        result = []
        for r in rows:
            returned = int(ret_map.get(r.id, 0) or 0)
            qty = r.quantity or 0
            # Net the money by the units returned (e.g. 5 sold, 1 returned → 4).
            if returned > 0:
                net_qty = max(qty - returned, 0)
                sp = r.selling_price or 0
                uc = r.unit_cost or 0
                dpct = r.discount_pct or 0
                net_subtotal = round(sp * net_qty, 2)
                net_total = round(net_subtotal - net_subtotal * dpct / 100, 2)
                net_profit = round(net_total - uc * net_qty, 2)
            else:
                net_total = r.total_amount or 0
                net_profit = r.gross_profit
            # Tax-inclusive total (matches receipt); profit stays ex-tax.
            net_total = round(net_total * tax_factor.get(r.id, 1.0), 2)
            result.append({
                "id": r.id, "part_id": r.part_id,
                "part_name": r.part.name if r.part else "—",
                "quantity": r.quantity, "reason": r.reason,
                "returned": returned,
                "job_ref": r.job_ref, "selling_price": r.selling_price,
                "discount_pct": r.discount_pct, "total_amount": r.total_amount,
                "gross_profit": r.gross_profit,
                "net_total": net_total, "net_profit": net_profit,
                "issued_by": r.issued_by, "issued_at": r.issued_at,
            })
        return ok(result)
    finally:
        db.close()


@app.route("/api/stock/out", methods=["POST"])
@login_required
def issue_stock():
    data = request.get_json() or {}
    data.setdefault("issued_by", current_user())
    db = get_session()
    try:
        schema = StockOutCreate(**data)
        txn = StockService(db).issue_stock(schema)
        DashboardService(db).invalidate()
        return ok({"id": txn.id, "quantity": txn.quantity,
                   "total_amount": txn.total_amount}), 201
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/stock/out/<int:txn_id>", methods=["DELETE"])
@admin_required
def cancel_stock_out(txn_id):
    db = get_session()
    try:
        StockService(db).cancel_stock_out(txn_id, user=current_user())
        DashboardService(db).invalidate()
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Suppliers ─────────────────────────────────────────────────────────────────


@app.route("/api/suppliers", methods=["GET"])
@login_required
def list_suppliers():
    include_inactive = request.args.get(
        "include_inactive", "false").lower() == "true"
    db = get_session()
    try:
        rows = SupplierService(db).get_all(include_inactive=include_inactive)
        return ok([{
            "id": s.id, "name": s.name, "contact_name": s.contact_name,
            "phone": s.phone, "email": s.email, "address": s.address,
            "notes": s.notes, "is_active": s.is_active,
        } for s in rows])
    finally:
        db.close()


@app.route("/api/suppliers", methods=["POST"])
@login_required
def create_supplier():
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = SupplierCreate(**data)
        supplier = SupplierService(db).create(schema)
        return ok({"id": supplier.id, "name": supplier.name}), 201
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/suppliers/<int:supplier_id>", methods=["PUT"])
@login_required
def update_supplier(supplier_id):
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = SupplierUpdate(**data)
        supplier = SupplierService(db).update(supplier_id, schema)
        return ok({"id": supplier.id, "name": supplier.name})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/suppliers/<int:supplier_id>", methods=["DELETE"])
@admin_required
def delete_supplier(supplier_id):
    db = get_session()
    try:
        SupplierService(db).delete(supplier_id)
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Dashboard ─────────────────────────────────────────────────────────────────


@app.route("/api/dashboard")
@login_required
def dashboard():
    db = get_session()
    try:
        m = DashboardService(db).get_metrics()
        return ok({
            "total_parts":       m.total_parts,
            "total_stock_value": m.total_stock_value,
            "low_stock_count":   m.low_stock_count,
            "stock_in_today":    m.stock_in_today,
            "stock_out_today":   m.stock_out_today,
            "sales_today":       m.sales_today,
            "profit_today":      m.profit_today,
            "sales_this_month":  m.sales_this_month,
            "profit_this_month": m.profit_this_month,
            "top_moving_parts":  m.top_moving_parts,
            "low_stock_parts":   m.low_stock_parts,
            "recent_activity":   m.recent_activity,
        })
    finally:
        db.close()


# ── Reports / Downloads ───────────────────────────────────────────────────────


def _auto_fit_columns(ws):
    from openpyxl.utils import get_column_letter

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            except Exception:
                pass

        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(
            ws.column_dimensions[letter].width or 0,
            min(max(max_len + 2, 11), 34)
        )


def _style_money_and_numbers(ws, header_row=5):
    """
    Apply number formats.

    Dashboard uses a Metric/Value layout, so the Value column must be
    formatted row-by-row. Counts and quantities must NOT show the peso sign.
    """
    if ws.title == "Dashboard":
        money_metrics = (
            "total stock value",
            "sales total",
            "gross profit",
        )
        number_metrics = (
            "total parts",
            "low stock count",
            "stock in transactions",
            "stock in quantity",
            "stock out transactions",
            "stock out quantity",
        )

        for row in range(header_row + 1, ws.max_row + 1):
            metric = str(ws.cell(row, 1).value or "").strip().lower()
            value_cell = ws.cell(row, 2)

            if any(name in metric for name in money_metrics):
                value_cell.number_format = '₱#,##0.00'
            elif any(name in metric for name in number_metrics):
                value_cell.number_format = '#,##0'
            else:
                value_cell.number_format = 'General'

        return

    money_words = ("cost", "price", "total", "profit", "amount")
    qty_words = ("quantity", "stock", "min", "parts", "count")

    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "").lower()

        if any(word in header for word in money_words):
            fmt = '₱#,##0.00'
        elif any(word in header for word in qty_words):
            fmt = '#,##0'
        elif "date" in header:
            fmt = 'yyyy-mm-dd hh:mm'
        else:
            fmt = None

        if fmt:
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row, col).number_format = fmt


def _write_sheet(wb, title, headers, rows, report_info=None):
    """
    Simple, clean Excel report sheet.
    No complicated merged layouts, so it is safer and easier to read on Excel/mobile.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from copy import copy as _copy

    ws = wb.create_sheet(title=title[:31])
    ws.sheet_view.showGridLines = False

    navy = "1D3461"
    orange = "1D3461"
    header_fill = "EAF1FB"
    light_fill = "F8FAFC"
    border_color = "D9E2EC"
    text_dark = "1E293B"
    muted = "64748B"
    green = "166534"
    red = "DC2626"

    # Build each style object ONCE and reuse the same instances for every cell.
    # openpyxl de-duplicates styles by value, so constructing a fresh
    # Font/Border/Fill/Alignment per cell (thousands of them on a big report)
    # was the bottleneck. Sharing instances keeps the style tables tiny and the
    # write an order of magnitude faster, with identical output.
    thin_side = Side(style="thin", color=border_color)
    cell_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)
    header_font_obj = Font(name="Segoe UI", size=10, bold=True, color=navy)
    header_fill_obj = PatternFill("solid", fgColor=header_fill)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font_obj = Font(name="Segoe UI", size=10, color=text_dark)
    low_font_obj = Font(name="Segoe UI", size=10, bold=True, color=red)
    ok_font_obj = Font(name="Segoe UI", size=10, bold=True, color=green)
    body_align = Alignment(vertical="center")
    body_fill_even = PatternFill("solid", fgColor=light_fill)
    body_fill_white = PatternFill("solid", fgColor="FFFFFF")

    max_cols = max(len(headers), 2)

    # Simple title + info rows
    # Keep "Inventra" as one word in one cell:
    # "Invent" = navy blue, "ra" = orange.
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont

        ws["A1"] = CellRichText(
            TextBlock(InlineFont(rFont="Segoe UI", sz=18,
                      b=True, color="FF1D3461"), "Invent"),
            TextBlock(InlineFont(rFont="Segoe UI", sz=18,
                      b=True, color="FFF59E0B"), "ra"),
        )
    except Exception:
        # Fallback if openpyxl rich text is unavailable.
        ws["A1"] = "Inventra"
        ws["A1"].font = Font(name="Segoe UI", size=18, bold=True, color=navy)

    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["B1"] = title
    ws["B1"].font = Font(name="Segoe UI", size=16, bold=True, color=text_dark)
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    # Make sure the full word and page title are visible.
    ws.column_dimensions["A"].width = max(
        ws.column_dimensions["A"].width or 0, 18)
    ws.column_dimensions["B"].width = max(
        ws.column_dimensions["B"].width or 0, 24)

    info = report_info or {}
    ws["A2"] = "Date Range"
    ws["B2"] = info.get("date_range", "All dates")
    ws["A3"] = "Generated"
    ws["B3"] = info.get("generated_at", "")
    ws["C3"] = "By"
    ws["D3"] = info.get("username", "system")

    for cell in ("A2", "A3", "C3"):
        ws[cell].font = Font(name="Segoe UI", size=9, bold=True, color=muted)
    for cell in ("B2", "B3", "D3"):
        ws[cell].font = Font(name="Segoe UI", size=9, color=text_dark)

    # Accent line
    for col in range(1, max_cols + 1):
        c = ws.cell(4, col)
        c.fill = PatternFill("solid", fgColor=orange)
        c.border = Border(bottom=Side(style="thin", color=orange))
    ws.row_dimensions[4].height = 4

    # Header row
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col)
        cell.value = header
        cell.font = header_font_obj
        cell.fill = header_fill_obj
        cell.alignment = header_align
        cell.border = cell_border
    ws.row_dimensions[header_row].height = 22

    # Data rows
    # Style one representative cell per (row-parity, font-kind), then COPY its
    # finished style array onto every other matching cell. Assigning styles via
    # the descriptors re-hashes the font/fill/border each time (the real cost on
    # big reports); copying the cached array is a cheap, hash-free array copy.
    style_templates = {}
    fonts_by_kind = {"low": low_font_obj, "ok": ok_font_obj, "": body_font_obj}

    for row_idx, row in enumerate(rows, start=header_row + 1):
        ws.append(row)
        parity = "e" if row_idx % 2 == 0 else "o"
        fill = body_fill_even if parity == "e" else body_fill_white

        for col in range(1, max_cols + 1):
            cell = ws.cell(row_idx, col)
            value = str(cell.value or "").upper()
            kind = "low" if value == "LOW" else "ok" if value == "OK" else ""
            key = (parity, kind)

            tmpl = style_templates.get(key)
            if tmpl is None:
                cell.fill = fill
                cell.alignment = body_align
                cell.border = cell_border
                cell.font = fonts_by_kind[kind]
                style_templates[key] = _copy(cell._style)
            else:
                cell._style = _copy(tmpl)

    if not rows:
        ws.append(["No records found"] + [""] * (max_cols - 1))
        cell = ws.cell(header_row + 1, 1)
        cell.font = Font(name="Segoe UI", size=10, italic=True, color=muted)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_cols)}{max(ws.max_row, header_row)}"

    _style_money_and_numbers(ws, header_row=header_row)
    _auto_fit_columns(ws)

    # Helpful fixed widths
    for col in range(1, max_cols + 1):
        header = str(ws.cell(header_row, col).value or "").lower()
        letter = get_column_letter(col)
        if "part" in header or "name" in header:
            ws.column_dimensions[letter].width = max(
                ws.column_dimensions[letter].width, 23)
        elif "date" in header:
            ws.column_dimensions[letter].width = max(
                ws.column_dimensions[letter].width, 19)
        elif "notes" in header or "reference" in header:
            ws.column_dimensions[letter].width = max(
                ws.column_dimensions[letter].width, 18)

    # Print-friendly
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws


def _parse_report_date(value, end_of_day=False):
    if not value:
        return None

    from datetime import datetime, time

    value = str(value).strip()
    try:
        if len(value) == 10:
            d = datetime.strptime(value, "%Y-%m-%d").date()
            return datetime.combine(d, time.max if end_of_day else time.min)
        return datetime.fromisoformat(value.replace("Z", "").replace("T", " "))
    except Exception:
        raise ValueError(f"Invalid date: {value}")


def _as_datetime(value):
    if value is None:
        return None

    from datetime import datetime, date, time

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text_value = str(value).strip()
    if not text_value:
        return None

    try:
        return datetime.fromisoformat(text_value.replace("Z", "").replace("T", " "))
    except Exception:
        try:
            return datetime.strptime(text_value[:10], "%Y-%m-%d")
        except Exception:
            return None


def _in_report_range(value, start_dt=None, end_dt=None):
    dt = _as_datetime(value)

    if start_dt and (dt is None or dt < start_dt):
        return False

    if end_dt and (dt is None or dt > end_dt):
        return False

    return True


def _date_range_label(start_date=None, end_date=None):
    if start_date and end_date:
        return f"{start_date} to {end_date}"
    if start_date:
        return f"From {start_date}"
    if end_date:
        return f"Until {end_date}"
    return "All dates"


def _make_xlsx_report(report_type, username="system", start_date=None, end_date=None):
    from datetime import datetime
    from openpyxl import Workbook

    os.makedirs(_REPORTS_DIR, exist_ok=True)

    clean_type = (report_type or "full").strip().lower()
    allowed = {"full", "inventory", "stock_in", "stock_out", "dashboard",
               "adjustments", "returns"}
    if clean_type not in allowed:
        clean_type = "full"

    start_dt = _parse_report_date(
        start_date, end_of_day=False) if start_date else None
    end_dt = _parse_report_date(
        end_date, end_of_day=True) if end_date else None

    if start_dt and end_dt and start_dt > end_dt:
        raise ValueError("From Date cannot be later than To Date.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inventra_{clean_type}_report_{timestamp}.xlsx"
    path = os.path.join(_REPORTS_DIR, filename)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    range_label = _date_range_label(start_date, end_date)
    report_info = {
        "date_range": range_label,
        "generated_at": generated_at,
        "username": username,
    }

    db = get_session()
    try:
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        stock_in_all = StockService(db).get_stock_in_history(limit=5000)
        stock_out_all = StockService(db).get_stock_out_history(limit=5000)

        stock_in_rows = [
            t for t in stock_in_all
            if _in_report_range(getattr(t, "received_at", None), start_dt, end_dt)
        ]
        stock_out_rows = [
            t for t in stock_out_all
            if _in_report_range(getattr(t, "issued_at", None), start_dt, end_dt)
        ]
        from database.models.sale import Sale
        sale_rows = [
            s for s in db.query(Sale).all()
            if _in_report_range(getattr(s, "sale_date", None), start_dt, end_dt)
        ]
        total_fees = round(sum((getattr(s, "labor_amount", 0) or 0)
                               for s in sale_rows), 2)

        if clean_type in ("full", "dashboard"):
            metrics = DashboardService(db).get_metrics()

            total_stock_in_qty = sum(
                (getattr(t, "quantity", 0) or 0) for t in stock_in_rows)
            total_stock_out_qty = sum(
                (getattr(t, "quantity", 0) or 0) for t in stock_out_rows)
            stock_out_sales = sum((getattr(t, "total_amount", 0) or 0)
                                  for t in stock_out_rows)
            stock_out_profit = sum((getattr(t, "gross_profit", 0) or 0)
                                   for t in stock_out_rows)
            total_sales = round(stock_out_sales + total_fees, 2)
            total_profit = round(stock_out_profit + total_fees, 2)

            rows = [
                ["Date Range", range_label],
                ["Total Parts", metrics.total_parts],
                ["Total Stock Value", metrics.total_stock_value],
                ["Low Stock Count", metrics.low_stock_count],
                ["Stock In Transactions", len(stock_in_rows)],
                ["Stock In Quantity", total_stock_in_qty],
                ["Stock Out Transactions", len(stock_out_rows)],
                ["Stock Out Quantity", total_stock_out_qty],
                ["Sales Total", total_sales],
                ["Gross Profit", total_profit],
            ]
            _write_sheet(wb, "Dashboard", [
                         "Metric", "Value"], rows, report_info=report_info)

            low_rows = []
            for item in DashboardService(db).get_low_stock_parts() or []:
                low_rows.append([
                    item.get("name", ""),
                    item.get("bin", ""),
                    item.get("current", ""),
                    item.get("min", ""),
                ])
            _write_sheet(
                wb,
                "Low Stock",
                ["Part", "Bin", "Current Stock", "Min Stock"],
                low_rows,
                report_info=report_info,
            )

        if clean_type in ("full", "inventory"):
            parts = PartsService(db).get_stock_view()
            rows = []
            for p in parts:
                rows.append([
                    p.get("sku", ""),
                    p.get("name", ""),
                    p.get("category", ""),
                    p.get("current_stock", ""),
                    p.get("unit", ""),
                    p.get("min_stock", ""),
                    p.get("bin_location", ""),
                    p.get("unit_cost", ""),
                    p.get("selling_price", ""),
                    "LOW" if p.get("is_low_stock") else "OK",
                ])
            _write_sheet(
                wb,
                "Inventory",
                ["SKU", "Part Name", "Category", "Stock", "Unit",
                    "Min", "Bin", "Cost", "Selling Price", "Status"],
                rows,
                report_info=report_info,
            )

        if clean_type in ("full", "stock_in"):
            rows = []
            for t in stock_in_rows:
                rows.append([
                    t.id,
                    str(t.received_at) if t.received_at else "",
                    t.part.name if t.part else "",
                    t.quantity,
                    t.unit_cost,
                    t.supplier.name if t.supplier else "",
                    t.reference_no or "",
                    t.notes or "",
                    t.received_by or "",
                ])
            _write_sheet(
                wb,
                "Stock In",
                ["Txn ID", "Date", "Part", "Quantity", "Unit Cost",
                    "Supplier", "Reference", "Notes", "Received By"],
                rows,
                report_info=report_info,
            )

        if clean_type in ("full", "stock_out"):
            rows = []
            for t in stock_out_rows:
                rows.append([
                    t.id,
                    str(t.issued_at) if t.issued_at else "",
                    t.part.name if t.part else "",
                    t.quantity,
                    t.selling_price,
                    t.discount_pct,
                    t.total_amount,
                    t.gross_profit,
                    t.reason or "",
                    t.job_ref or "",
                    t.issued_by or "",
                ])
            stock_out_qty = sum((t.quantity or 0) for t in stock_out_rows)
            stock_out_total = round(sum((t.total_amount or 0)
                                        for t in stock_out_rows), 2)
            stock_out_profit = round(sum((t.gross_profit or 0)
                                         for t in stock_out_rows), 2)
            if stock_out_rows:
                rows.append([
                    "", "", "TOTAL",
                    stock_out_qty,
                    "", "",
                    stock_out_total,
                    stock_out_profit,
                    "", "", "",
                ])
            if total_fees:
                rows.append([
                    "", "", "FEES",
                    "", "", "",
                    total_fees,
                    total_fees,
                    "", "", "",
                ])
            if stock_out_rows or total_fees:
                rows.append([
                    "", "", "NET TOTAL",
                    stock_out_qty,
                    "", "",
                    round(stock_out_total + total_fees, 2),
                    round(stock_out_profit + total_fees, 2),
                    "", "", "",
                ])
            _write_sheet(
                wb,
                "Stock Out Sales",
                ["Txn ID", "Date", "Part", "Quantity", "Selling Price", "Discount %",
                    "Total", "Gross Profit", "Reason", "Job Ref", "Issued By"],
                rows,
                report_info=report_info,
            )

        if clean_type in ("full", "adjustments"):
            from core.services.adjustment_service import (
                AdjustmentService, REASONS as ADJ_REASONS)
            rows = []
            for r in AdjustmentService(db).get_history(
                    date_from=start_date, date_to=end_date, limit=10000):
                rows.append([
                    str(r.created_at)[:19].replace("T", " ") if r.created_at else "",
                    r.part.sku if r.part else "",
                    r.part.name if r.part else "",
                    ADJ_REASONS.get(r.reason_code, r.reason_code),
                    r.previous_count, r.new_count, r.delta, r.value_delta,
                    r.user, r.note or "",
                ])
            _write_sheet(
                wb, "Adjustments",
                ["Date", "SKU", "Part", "Reason", "Prev", "New",
                 "Delta Qty", "Delta Value", "User", "Note"],
                rows, report_info=report_info)

        if clean_type in ("full", "returns"):
            from core.services.return_service import (
                ReturnService, REASONS as RET_REASONS, CONDITIONS)
            rows = []
            for r in ReturnService(db).get_history(
                    date_from=start_date, date_to=end_date, limit=10000):
                rows.append([
                    str(r.created_at)[:19].replace("T", " ") if r.created_at else "",
                    r.part.sku if r.part else "",
                    r.part.name if r.part else "",
                    r.quantity, CONDITIONS.get(r.condition, r.condition),
                    RET_REASONS.get(r.reason_code, r.reason_code),
                    r.refund_amount, r.refund_method, r.restock_qty,
                    r.profit_delta, r.user,
                ])
            _write_sheet(
                wb, "Returns",
                ["Date", "SKU", "Part", "Qty", "Condition", "Reason",
                 "Refund", "Method", "Restocked", "Profit Impact", "By"],
                rows, report_info=report_info)

        wb.properties.title = "Inventra Report"
        wb.properties.subject = str(clean_type).replace("_", " ").title()
        wb.properties.creator = username

        wb.save(path)
        return filename, path
    finally:
        db.close()


@app.route("/api/reports/generate", methods=["POST"])
@login_required
def generate_report():
    data = request.get_json() or {}
    report_type = data.get("report_type", "full")
    start_date = data.get("start_date") or None
    end_date = data.get("end_date") or None

    try:
        filename, path = _make_xlsx_report(
            report_type,
            username=current_user(),
            start_date=start_date,
            end_date=end_date,
        )
        return ok({
            "filename": filename,
            "download_url": f"/api/reports/download/{filename}",
            "size": os.path.getsize(path),
        })
    except ImportError:
        return err("openpyxl is required to generate Excel reports. Install it with: pip install openpyxl")
    except Exception as e:
        return err(str(e))


@app.route("/api/reports", methods=["GET"])
@login_required
def list_reports():
    """
    List generated report files from the export/report folder.
    """
    try:
        os.makedirs(_REPORTS_DIR, exist_ok=True)
        files = []

        for name in os.listdir(_REPORTS_DIR):
            if name.startswith("~$"):
                continue

            path = os.path.join(_REPORTS_DIR, name)
            if not os.path.isfile(path):
                continue

            ext = os.path.splitext(name)[1].lower()
            if ext not in REPORT_EXTENSIONS:
                continue

            stat = os.stat(path)
            files.append({
                "name": name,
                "ext": ext.lstrip("."),
                "size": stat.st_size,
                "modified_at": __import__("datetime").datetime.fromtimestamp(
                    stat.st_mtime
                ).isoformat(timespec="seconds"),
            })

        files.sort(key=lambda item: item["modified_at"], reverse=True)
        return ok(files)
    except Exception as e:
        return err(str(e))


@app.route("/api/reports/download/<path:filename>", methods=["GET"])
@login_required
def download_report(filename):
    """
    Download a generated report file.
    """
    safe_name = os.path.basename(filename)
    ext = os.path.splitext(safe_name)[1].lower()

    if not safe_name or ext not in REPORT_EXTENSIONS:
        return err("Invalid report file", 400)

    full_path = os.path.abspath(os.path.join(_REPORTS_DIR, safe_name))
    if not full_path.startswith(_REPORTS_DIR):
        return err("Invalid report path", 400)

    if not os.path.exists(full_path):
        return err("Report not found", 404)

    return send_from_directory(_REPORTS_DIR, safe_name, as_attachment=True)


@app.route("/api/reports/<path:filename>", methods=["DELETE"])
@admin_required
def delete_report(filename):
    """
    Delete a generated report file.
    """
    safe_name = os.path.basename(filename)
    ext = os.path.splitext(safe_name)[1].lower()

    if not safe_name or ext not in REPORT_EXTENSIONS:
        return err("Invalid report file", 400)

    full_path = os.path.abspath(os.path.join(_REPORTS_DIR, safe_name))
    if not full_path.startswith(_REPORTS_DIR):
        return err("Invalid report path", 400)

    if not os.path.exists(full_path):
        return err("Report not found", 404)

    try:
        os.remove(full_path)
        return ok()
    except PermissionError:
        return err("Close the report file and try again.", 409)
    except Exception as e:
        return err(str(e))

# ── POS ───────────────────────────────────────────────────────────────────────


@app.route("/api/pos/parts", methods=["GET"])
@login_required
def pos_parts():
    search = request.args.get("search", "")
    category = request.args.get("category", "")
    db = get_session()
    try:
        rows = PosService(db).search_parts(search, category=category, limit=200)
        return ok(rows)
    finally:
        db.close()


@app.route("/api/pos/sales", methods=["POST"])
@login_required
def pos_create_sale():
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = SaleCreate(**data)
        sale = PosService(db).create_sale(schema, cashier=current_display_name())
        detail = PosService(db).get_sale_detail(sale.id)
        DashboardService(db).invalidate()
        return ok(detail), 201
    except ValueError as e:
        db.rollback()
        return err(str(e))
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/pos/sales", methods=["GET"])
@login_required
def pos_recent_sales():
    date_from = request.args.get("date_from") or None
    date_to = request.args.get("date_to") or None
    db = get_session()
    try:
        sales = PosService(db).get_recent_sales(limit=200, date_from=date_from,
                                                date_to=date_to)
        return ok([{
            "id": s.id, "receipt_no": s.receipt_no, "sale_date": s.sale_date,
            "cashier": s.cashier, "payment_method": s.payment_method,
            "item_count": sum(i.quantity for i in s.items), "grand_total": s.grand_total,
        } for s in sales])
    finally:
        db.close()


@app.route("/api/pos/sales/<int:sale_id>", methods=["GET"])
@login_required
def pos_sale_detail(sale_id):
    db = get_session()
    try:
        detail = PosService(db).get_sale_detail(sale_id)
        if not detail:
            return err("Sale not found", 404)
        return ok(detail)
    finally:
        db.close()


@app.route("/api/pos/sales/<int:sale_id>/receipt", methods=["GET"])
@login_required
def pos_sale_receipt(sale_id):
    """Server-rendered printable receipt HTML (reprint never mutates data)."""
    db = get_session()
    try:
        detail = PosService(db).get_sale_detail(sale_id)
        if not detail:
            return err("Sale not found", 404)
        return render_receipt_html(detail), 200, {"Content-Type": "text/html; charset=utf-8"}
    finally:
        db.close()


@app.route("/api/pos/sales/<int:sale_id>/void", methods=["POST"])
@login_required
def pos_void_sale(sale_id):
    db = get_session()
    try:
        receipt = PosService(db).void_sale(sale_id, user=current_user())
        DashboardService(db).invalidate()
        return ok({"receipt_no": receipt})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/pos/fee-types", methods=["GET"])
@login_required
def pos_get_fee_types():
    db = get_session()
    try:
        return ok(SettingsService(db).get_fee_types())
    finally:
        db.close()


@app.route("/api/pos/fee-types", methods=["PUT"])
@admin_required
def pos_set_fee_types():
    data = request.get_json() or {}
    db = get_session()
    try:
        saved = SettingsService(db).set_fee_types(data.get("fee_types", []))
        return ok(saved)
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/pos/settings", methods=["GET"])
@login_required
def pos_get_settings():
    db = get_session()
    try:
        return ok(SettingsService(db).get_pos_settings())
    finally:
        db.close()


@app.route("/api/pos/settings", methods=["PUT"])
@admin_required
def pos_update_settings():
    data = request.get_json() or {}
    db = get_session()
    try:
        schema = PosSettingsUpdate(**data)
        updated = SettingsService(db).update_pos_settings(
            schema.model_dump(exclude_none=True))
        return ok(updated)
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Users (admin only) ────────────────────────────────────────────────────────


@app.route("/api/users", methods=["GET"])
@admin_required
def list_users():
    db = get_session()
    try:
        users = AuthService(db).get_all_users()
        return ok([{
            "id": u.id, "username": u.username, "full_name": u.full_name,
            "role": u.role, "is_active": u.is_active,
        } for u in users])
    finally:
        db.close()


@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    data = request.get_json() or {}
    db = get_session()
    try:
        user = AuthService(db).create_user(
            username=data.get("username", ""),
            full_name=data.get("full_name", ""),
            role=data.get("role", "staff"),
            password=data.get("password", ""),
        )
        return ok({"id": user.id, "username": user.username}), 201
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    data = request.get_json() or {}
    db = get_session()
    try:
        AuthService(db).update_user(
            user_id,
            full_name=data.get("full_name"),
            role=data.get("role"),
            password=data.get("password"),
        )
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/users/<int:user_id>/deactivate", methods=["POST"])
@admin_required
def deactivate_user(user_id):
    db = get_session()
    try:
        AuthService(db).deactivate_user(user_id)
        return ok()
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()

# ── Startup ───────────────────────────────────────────────────────────────────


def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


if __name__ == "__main__":
    print("🔧 Initialising database...")
    init_db()
    ip = get_local_ip()
    port = 5000
    print(f"\n✅ Inventra Web is running!")
    print(f"   Local:    http://localhost:{port}")
    print(f"   Network:  http://{ip}:{port}   ← use this on Android")
    print(f"\n   Default login: admin / admin123")
    print(f"   Press Ctrl+C to stop.\n")
    app.run(host="0.0.0.0", port=port, debug=False)

# ── These replace the old DELETE-only routes above with full CRUD ──────────────
# (The original routes stay for backward compat; these add GET + PUT)


@app.route("/api/stock/in/<int:txn_id>/detail", methods=["GET"])
@login_required
def get_stock_in_txn(txn_id):
    from database.models.stock_in import StockIn
    db = get_session()
    try:
        t = db.get(StockIn, txn_id)
        if not t:
            return err("Transaction not found", 404)
        return ok({"id": t.id, "part_id": t.part_id,
                   "part_name": t.part.name if t.part else "—",
                   "supplier_id": t.supplier_id,
                   "quantity": t.quantity, "unit_cost": t.unit_cost,
                   "reference_no": t.reference_no,
                   "received_at": str(t.received_at) if t.received_at else None})
    finally:
        db.close()


@app.route("/api/stock/in/<int:txn_id>/update", methods=["POST"])
@login_required
def update_stock_in_txn(txn_id):
    data = request.get_json() or {}
    db = get_session()
    try:
        new_qty = int(data.get("quantity", 0))
        if new_qty <= 0:
            return err("Quantity must be greater than 0")

        txn = StockService(db).update_stock_in(
            txn_id=txn_id,
            new_quantity=new_qty,
            new_unit_cost=float(data.get("unit_cost") or 0),
            new_supplier_id=data.get("supplier_id") or None,
            new_reference_no=data.get("reference_no") or None,
            user=current_user(),
        )

        DashboardService(db).invalidate()
        return ok({"id": txn.id})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()


@app.route("/api/stock/out/<int:txn_id>/detail", methods=["GET"])
@login_required
def get_stock_out_txn(txn_id):
    from database.models.stock_out import StockOut
    db = get_session()
    try:
        t = db.get(StockOut, txn_id)
        if not t:
            return err("Transaction not found", 404)
        return ok({"id": t.id, "part_id": t.part_id,
                   "part_name": t.part.name if t.part else "—",
                   "quantity": t.quantity, "selling_price": t.selling_price,
                   "discount_pct": t.discount_pct, "reason": t.reason,
                   "job_ref": t.job_ref, "total_amount": t.total_amount,
                   "gross_profit": t.gross_profit,
                   "issued_at": str(t.issued_at) if t.issued_at else None})
    finally:
        db.close()


@app.route("/api/stock/out/<int:txn_id>/update", methods=["POST"])
@login_required
def update_stock_out_txn(txn_id):
    data = request.get_json() or {}
    db = get_session()
    try:
        new_qty = int(data.get("quantity", 0))
        if new_qty <= 0:
            return err("Quantity must be greater than 0")

        txn = StockService(db).update_stock_out(
            txn_id=txn_id,
            new_quantity=new_qty,
            new_price=float(data.get("selling_price") or 0),
            new_disc_pct=float(data.get("discount_pct") or 0),
            new_reason=data.get("reason") or "Other",
            new_job_ref=data.get("job_ref") or None,
            user=current_user(),
        )

        DashboardService(db).invalidate()
        return ok({"id": txn.id})
    except Exception as e:
        db.rollback()
        return err(str(e))
    finally:
        db.close()
