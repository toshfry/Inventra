from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import List
import os
from datetime import datetime
from config.settings import EXPORT_DIR


class ReportService:

    def __init__(self, db: Session):
        self.db = db

    def _matches_report_date(self, row: dict, date_from: str = None, date_to: str = None) -> bool:
        """
        Used for inventory/category exports.
        A part is included when its Last Received or Last Issued date is inside
        the selected report date range.
        """
        if not date_from and not date_to:
            return True

        dates = [row.get("last_received"), row.get("last_issued")]
        dates = [d for d in dates if d and d != "Never"]

        if not dates:
            return False

        for d in dates:
            if date_from and d < date_from:
                continue
            if date_to and d > date_to:
                continue
            return True

        return False

    def _filter_inventory_rows(self, rows: List[dict], date_from: str = None, date_to: str = None) -> List[dict]:
        """
        Filter inventory rows by selected report date.
        """
        return [r for r in rows if self._matches_report_date(r, date_from, date_to)]

    def _category_from_inventory_rows(self, rows: List[dict]) -> List[dict]:
        """
        Build the By Category report from already-filtered inventory rows.
        """
        grouped = {}

        for r in rows:
            category = r.get("category") or "Uncategorized"
            item = grouped.setdefault(category, {
                "category": category,
                "parts": 0,
                "total_stock": 0,
                "value": 0,
                "low_count": 0,
            })

            item["parts"] += 1
            item["total_stock"] += r.get("current_stock", 0) or 0
            item["value"] += r.get("stock_value", 0) or 0

            if r.get("is_low_stock"):
                item["low_count"] += 1

        return sorted(grouped.values(), key=lambda x: x["value"], reverse=True)

    # ── Full part-level inventory detail ─────────────────────────────
    def inventory_detail(self) -> List[dict]:
        """Every active part with full stock, cost, and movement data."""
        rows = self.db.execute(text("""
            SELECT
                ps.sku,
                ps.name,
                ps.category,
                ps.current_stock,
                ps.min_stock,
                ps.unit,
                ps.unit_cost,
                ps.stock_value,
                ps.bin_location,
                ps.total_in,
                ps.total_out,
                ps.is_low_stock,
                ps.vehicle_makes,
                (SELECT MAX(si.received_at)
                 FROM stock_in si WHERE si.part_id = ps.id) AS last_received,
                (SELECT MAX(so.issued_at)
                 FROM stock_out so WHERE so.part_id = ps.id) AS last_issued,
                (SELECT s.name FROM stock_in si
                 JOIN suppliers s ON s.id = si.supplier_id
                 WHERE si.part_id = ps.id
                 ORDER BY si.received_at DESC LIMIT 1) AS last_supplier
            FROM part_stock ps
            ORDER BY ps.category, ps.name
        """)).fetchall()
        return [
            {
                "sku":           r[0],
                "name":          r[1],
                "category":      r[2] or "Uncategorized",
                "current_stock": r[3] or 0,
                "min_stock":     r[4] or 0,
                "unit":          r[5] or "pcs",
                "unit_cost":     round(r[6] or 0, 2),
                "stock_value":   round(r[7] or 0, 2),
                "bin_location":  r[8] or "—",
                "total_in":      r[9] or 0,
                "total_out":     r[10] or 0,
                "is_low_stock":  bool(r[11]),
                "vehicle_makes": r[12] or "—",
                "last_received": (r[13] or "")[:10] or "Never",
                "last_issued":   (r[14] or "")[:10] or "Never",
                "last_supplier": r[15] or "—",
                "status":        "LOW STOCK" if r[11] else "OK",
            }
            for r in rows
        ]

    # ── Category summary ─────────────────────────────────────────────
    def inventory_by_category(self) -> List[dict]:
        rows = self.db.execute(text("""
            SELECT
                category,
                COUNT(*)            AS parts,
                SUM(current_stock)  AS total_stock,
                SUM(stock_value)    AS value,
                SUM(is_low_stock)   AS low_count
            FROM part_stock
            GROUP BY category
            ORDER BY value DESC
        """)).fetchall()
        return [
            {"category": r[0] or "Uncategorized", "parts": r[1],
             "total_stock": r[2] or 0, "value": round(r[3] or 0, 2),
             "low_count": r[4] or 0}
            for r in rows
        ]

    # ── Movement history ─────────────────────────────────────────────
    def movement_history(self, date_from: str = None, date_to: str = None,
                         part_id: int = None) -> List[dict]:
        params = {}
        sql = """
            SELECT 'IN' AS type, si.received_at AS ts,
                   p.name, p.sku, p.unit,
                   si.quantity, si.unit_cost,
                   COALESCE(s.name, '—') AS supplier,
                   si.received_by, si.reference_no, si.notes
            FROM stock_in si
            JOIN parts p ON p.id = si.part_id
            LEFT JOIN suppliers s ON s.id = si.supplier_id
            WHERE 1=1
        """
        if part_id:
            sql += " AND si.part_id=:pid"
            params["pid"] = part_id
        if date_from:
            sql += " AND DATE(si.received_at)>=:df"
            params["df"] = date_from
        if date_to:
            sql += " AND DATE(si.received_at)<=:dt"
            params["dt"] = date_to

        sql += """
            UNION ALL
            SELECT 'OUT', so.issued_at,
                   p.name, p.sku, p.unit,
                   so.quantity, 0,
                   so.reason, so.issued_by,
                   COALESCE(so.job_ref, '—'), '—'
            FROM stock_out so
            JOIN parts p ON p.id = so.part_id
            WHERE 1=1
        """
        if part_id:
            sql += " AND so.part_id=:pid"
        if date_from:
            sql += " AND DATE(so.issued_at)>=:df"
        if date_to:
            sql += " AND DATE(so.issued_at)<=:dt"
        sql += " ORDER BY ts DESC LIMIT 1000"

        rows = self.db.execute(text(sql), params).fetchall()
        return [
            {
                "type":      r[0],
                "timestamp": r[1],
                "part":      r[2],
                "sku":       r[3],
                "unit":      r[4] or "pcs",
                "quantity":  r[5],
                "unit_cost": round(r[6] or 0, 2),
                "supplier":  r[7],
                "user":      r[8],
                "reference": r[9] or "—",
                "notes":     r[10] or "—",
            }
            for r in rows
        ]

    # ── Aging analysis ───────────────────────────────────────────────
    def aging_analysis(self, days: int = 90) -> List[dict]:
        rows = self.db.execute(text(f"""
            SELECT ps.name, ps.sku, ps.category,
                   ps.current_stock, ps.unit, ps.unit_cost, ps.stock_value,
                   ps.bin_location,
                   MAX(so.issued_at) AS last_issued,
                   ps.total_in, ps.total_out
            FROM part_stock ps
            LEFT JOIN stock_out so ON so.part_id = ps.id
            GROUP BY ps.id
            HAVING last_issued IS NULL
               OR DATE(last_issued) <= DATE('now', '-{days} days')
            ORDER BY last_issued ASC NULLS FIRST
        """)).fetchall()
        return [
            {
                "name":        r[0], "sku": r[1],
                "category":    r[2] or "—",
                "stock":       r[3] or 0,
                "unit":        r[4] or "pcs",
                "unit_cost":   round(r[5] or 0, 2),
                "value":       round(r[6] or 0, 2),
                "bin":         r[7] or "—",
                "last_issued": (r[8] or "")[:10] or "Never",
                "total_in":    r[9] or 0,
                "total_out":   r[10] or 0,
            }
            for r in rows
        ]

    # ── Summary stats ────────────────────────────────────────────────
    def summary_stats(self) -> dict:
        row = self.db.execute(text("""
            SELECT
                COUNT(*)              AS total_parts,
                SUM(current_stock)    AS total_units,
                SUM(stock_value)      AS total_value,
                SUM(is_low_stock)     AS low_stock,
                SUM(CASE WHEN current_stock = 0 THEN 1 ELSE 0 END) AS zero_stock
            FROM part_stock
        """)).fetchone()
        return {
            "total_parts":  row[0] or 0,
            "total_units":  row[1] or 0,
            "total_value":  round(row[2] or 0, 2),
            "low_stock":    row[3] or 0,
            "zero_stock":   row[4] or 0,
        }

    # ── Sales report ─────────────────────────────────────────────────
    def sales_detail(self, date_from: str = None, date_to: str = None) -> List[dict]:
        params = {}
        sql = """
            SELECT
                so.issued_at,
                p.sku,
                p.name,
                c.name          AS category,
                so.quantity,
                p.unit,
                so.unit_cost,
                so.selling_price,
                so.discount_pct,
                so.discount_amount,
                so.subtotal,
                so.total_amount,
                so.gross_profit,
                so.reason,
                so.job_ref,
                so.issued_by
            FROM stock_out so
            JOIN parts p ON p.id = so.part_id
            LEFT JOIN categories c ON c.id = p.category_id
            WHERE so.reason IN ('Sale', 'POS Sale')
        """
        if date_from:
            sql += " AND DATE(so.issued_at) >= :df"
            params["df"] = date_from
        if date_to:
            sql += " AND DATE(so.issued_at) <= :dt"
            params["dt"] = date_to
        sql += " ORDER BY so.issued_at DESC LIMIT 2000"

        rows = self.db.execute(text(sql), params).fetchall()
        return [
            {
                "issued_at":       (r[0] or "")[:19].replace("T", " "),
                "sku":             r[1],
                "name":            r[2],
                "category":        r[3] or "—",
                "quantity":        r[4],
                "unit":            r[5] or "pcs",
                "unit_cost":       round(r[6] or 0, 2),
                "selling_price":   round(r[7] or 0, 2),
                "discount_pct":    round(r[8] or 0, 2),
                "discount_amount": round(r[9] or 0, 2),
                "subtotal":        round(r[10] or 0, 2),
                "total_amount":    round(r[11] or 0, 2),
                "gross_profit":    round(r[12] or 0, 2),
                "reason":          r[13] or "—",
                "job_ref":         r[14] or "—",
                "issued_by":       r[15] or "—",
                "margin_pct":      round(
                    ((r[11] or 0) - (r[6] or 0) * (r[4] or 0))
                    / max((r[6] or 0) * (r[4] or 0), 1) * 100, 1
                ),
            }
            for r in rows
        ]

    def sales_summary_stats(self, date_from: str = None, date_to: str = None) -> dict:
        params = {}
        # Only genuine sales count as revenue/profit — exclude Job Use,
        # Damaged/Write-off, Adjustment, Customer Return, Other.
        where = "WHERE reason IN ('Sale', 'POS Sale')"
        if date_from:
            where += " AND DATE(issued_at) >= :df"
            params["df"] = date_from
        if date_to:
            where += " AND DATE(issued_at) <= :dt"
            params["dt"] = date_to

        row = self.db.execute(text(f"""
            SELECT
                COUNT(*)                        AS transactions,
                COALESCE(SUM(quantity),0)        AS units_sold,
                COALESCE(SUM(subtotal),0)        AS gross_revenue,
                COALESCE(SUM(discount_amount),0) AS total_discounts,
                COALESCE(SUM(total_amount),0)    AS net_revenue,
                COALESCE(SUM(gross_profit),0)    AS gross_profit,
                COALESCE(SUM(unit_cost*quantity),0) AS total_cost
            FROM stock_out {where}
        """), params).fetchone()
        net = row[4] or 0
        cost = row[6] or 0
        gross_profit = row[5] or 0

        # Net out customer returns processed in the same date range.
        rwhere = "WHERE 1=1"
        if date_from:
            rwhere += " AND DATE(created_at) >= :df"
        if date_to:
            rwhere += " AND DATE(created_at) <= :dt"
        rrow = self.db.execute(text(
            f"SELECT COALESCE(SUM(refund_amount),0), COALESCE(SUM(profit_delta),0), "
            f"COUNT(*) FROM customer_returns {rwhere}"), params).fetchone()
        refunds = round(rrow[0] or 0, 2)
        ret_profit_delta = round(rrow[1] or 0, 2)
        ret_count = rrow[2] or 0

        # Labor / service fees are extra revenue with no cost → add to both
        # revenue and profit. Keyed by the sale date.
        lwhere = "WHERE 1=1"
        if date_from:
            lwhere += " AND DATE(sale_date) >= :df"
        if date_to:
            lwhere += " AND DATE(sale_date) <= :dt"
        labor = round(self.db.execute(text(
            f"SELECT COALESCE(SUM(labor_amount),0) FROM sales {lwhere}"),
            params).scalar() or 0, 2)

        net_revenue = round(net - refunds + labor, 2)
        gross_profit = round(gross_profit + ret_profit_delta + labor, 2)
        return {
            "transactions":   row[0] or 0,
            "units_sold":     row[1] or 0,
            "gross_revenue":  round(row[2] or 0, 2),
            "total_discounts": round(row[3] or 0, 2),
            "refunds":        refunds,
            "returns_count":  ret_count,
            "labor":          labor,
            "net_revenue":    net_revenue,
            "gross_profit":   gross_profit,
            "total_cost":     round(cost, 2),
            "margin_pct":     round(gross_profit / max(net_revenue, 1) * 100, 1) if net_revenue else 0.0,
        }

    def sales_by_part(self, date_from: str = None, date_to: str = None) -> List[dict]:
        params = {}
        where = "WHERE so.reason IN ('Sale', 'POS Sale')"
        if date_from:
            where += " AND DATE(so.issued_at) >= :df"
            params["df"] = date_from
        if date_to:
            where += " AND DATE(so.issued_at) <= :dt"
            params["dt"] = date_to

        rows = self.db.execute(text(f"""
            SELECT
                p.sku, p.name,
                c.name          AS category,
                SUM(so.quantity)               AS qty_sold,
                p.unit,
                AVG(so.selling_price)          AS avg_price,
                SUM(so.total_amount)           AS revenue,
                SUM(so.gross_profit)           AS profit,
                COUNT(so.id)                   AS txn_count
            FROM stock_out so
            JOIN parts p ON p.id = so.part_id
            LEFT JOIN categories c ON c.id = p.category_id
            {where}
            GROUP BY p.id
            ORDER BY revenue DESC
        """), params).fetchall()
        return [
            {
                "sku":       r[0], "name": r[1], "category": r[2] or "—",
                "qty_sold":  r[3] or 0, "unit": r[4] or "pcs",
                "avg_price": round(r[5] or 0, 2),
                "revenue":   round(r[6] or 0, 2),
                "profit":    round(r[7] or 0, 2),
                "txn_count": r[8] or 0,
                "margin_pct": round((r[7] or 0) / max(r[6] or 1, 1) * 100, 1),
            }
            for r in rows
        ]

    # ── Audit log (Added to supply data for the export) ──────────────
    def audit_log(self, date_from: str = None, date_to: str = None) -> List[dict]:
        """
        Fetch Audit Log data for the Reports export.

        Inventra's real audit model/table is audit_log with:
        created_at, user, action, reason, delta, and part relationship.
        """
        try:
            from database.models.audit_log import AuditLog

            q = self.db.query(AuditLog)

            if date_from:
                q = q.filter(AuditLog.created_at >= date_from)
            if date_to:
                q = q.filter(AuditLog.created_at <= date_to + "T23:59:59")

            logs = q.order_by(AuditLog.created_at.desc()).limit(2000).all()

            rows = []
            for log in logs:
                part_text = ""
                if getattr(log, "part", None):
                    sku = getattr(log.part, "sku", "") or ""
                    name = getattr(log.part, "name", "") or ""
                    part_text = f"{sku} {name}".strip()

                reason = log.reason or ""
                if log.delta is not None:
                    delta = f"+{log.delta}" if log.delta > 0 else str(
                        log.delta)
                    reason = f"{reason} | Change: {delta}" if reason else f"Change: {delta}"

                if part_text:
                    reason = f"{part_text} — {reason}" if reason else part_text

                rows.append({
                    "timestamp":   (log.created_at or "")[:19].replace("T", " "),
                    "username":    log.user or "—",
                    "action":      log.action or "—",
                    "module":      "Inventory",
                    "description": reason or "—",
                })

            return rows
        except Exception:
            return []

    # ── XLSX-only export compatibility layer ──────────────────────────
    def export_to_csv(self, report_type: str = "inventory", date_from: str = None, date_to: str = None) -> str:
        """
        Backward-compatible method name.

        Older UI code may still call export_to_csv(...). Inventra now exports
        Excel files only, so this method redirects to export_to_excel(...) and
        returns an .xlsx file path.
        """
        return self.export_to_excel(report_type, date_from=date_from, date_to=date_to)

    def export_sales_to_csv(self, date_from: str = None, date_to: str = None) -> str:
        """Backward-compatible sales export. Always returns .xlsx."""
        return self.export_to_excel("sales", date_from=date_from, date_to=date_to)

    def export_audit_to_csv(self, date_from: str = None, date_to: str = None) -> str:
        """Backward-compatible audit export. Always returns .xlsx."""
        return self.export_to_excel("audit", date_from=date_from, date_to=date_to)

    def export_sales_csv(self, date_from: str = None, date_to: str = None) -> str:
        """Backward-compatible sales export. Always returns .xlsx."""
        return self.export_to_excel("sales", date_from=date_from, date_to=date_to)

    def export_audit_csv(self, date_from: str = None, date_to: str = None) -> str:
        """Backward-compatible audit export. Always returns .xlsx."""
        return self.export_to_excel("audit", date_from=date_from, date_to=date_to)

    def export_sales_excel(self, date_from: str = None, date_to: str = None) -> str:
        return self.export_to_excel("sales", date_from=date_from, date_to=date_to)

    def export_audit_excel(self, date_from: str = None, date_to: str = None) -> str:
        return self.export_to_excel("audit", date_from=date_from, date_to=date_to)

    # ── XLSX export with autofit ──────────────────────────────────────

    def export_to_excel(self, report_type: str = "inventory", date_from: str = None,
                        date_to: str = None, low_only: bool = False) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError(
                "openpyxl not installed. Run: pip install openpyxl")

        os.makedirs(EXPORT_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = "_lowstock" if (low_only and report_type == "inventory") else ""
        filename = os.path.join(
            EXPORT_DIR, f"inventra_{report_type}{suffix}_{ts}.xlsx")

        wb = Workbook()
        ws = wb.active

        # ── Header style ──────────────────────────────────────────────
        hdr_fill = PatternFill("solid", fgColor="1D3461")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0D0D0")
        cell_border = Border(bottom=Side(style="thin", color="E5E5E5"))

        alt_fill = PatternFill("solid", fgColor="F7F6F3")
        num_align = Alignment(horizontal="right")
        ctr_align = Alignment(horizontal="center")

        # ── Choose data ───────────────────────────────────────────────
        if report_type == "inventory":
            ws.title = "Low Stock" if low_only else "Parts Inventory"
            headers = ["SKU", "Part Name", "Category", "Stock", "Unit", "Min Stock",
                       "Status", "Unit Cost", "Stock Value",
                       "Total In", "Total Out", "Bin", "Last Supplier",
                       "Last Received", "Last Issued", "Vehicle Makes"]
            data = self._filter_inventory_rows(
                self.inventory_detail(), date_from, date_to)
            if low_only:
                data = [d for d in data if d.get("is_low_stock")]

            def row_values(d):
                return [d["sku"], d["name"], d["category"], d["current_stock"], d["unit"],
                        d["min_stock"], d["status"], d["unit_cost"],
                        d["stock_value"], d["total_in"], d["total_out"],
                        d["bin_location"], d["last_supplier"],
                        d["last_received"], d["last_issued"], d["vehicle_makes"]]
            num_cols = {8, 9}  # 1-based columns that are currency

        elif report_type == "movements":
            ws.title = "Movements"
            headers = ["Type", "Date/Time", "Part Name", "SKU", "Unit", "Qty",
                       "Unit Cost", "Supplier/Reason", "Reference", "User"]
            data = self.movement_history(date_from=date_from, date_to=date_to)

            def row_values(d):
                return [d["type"], d["timestamp"][:19].replace("T", " "),
                        d["part"], d["sku"], d["unit"], d["quantity"],
                        d["unit_cost"], d["supplier"], d["reference"], d["user"]]
            num_cols = {7}

        elif report_type == "aging":
            ws.title = "Aging Analysis"
            headers = ["Part Name", "SKU", "Category", "Stock", "Unit",
                       "Unit Cost", "Total Value", "Bin", "Last Issued",
                       "Total In", "Total Out"]
            data = self.aging_analysis()

            def row_values(d):
                return [d["name"], d["sku"], d["category"], d["stock"], d["unit"],
                        d["unit_cost"], d["value"], d["bin"], d["last_issued"],
                        d["total_in"], d["total_out"]]
            num_cols = {6, 7}

        elif report_type == "category":
            ws.title = "By Category"
            headers = ["Category", "Total Parts",
                       "Total Stock", "Stock Value", "Low Stock Items"]
            data = self._category_from_inventory_rows(
                self._filter_inventory_rows(
                    self.inventory_detail(), date_from, date_to)
            )

            def row_values(d):
                return [d["category"], d["parts"], d["total_stock"], d["value"], d["low_count"]]
            num_cols = {4}

        elif report_type == "sales":
            ws.title = "Sales"
            headers = ["Date/Time", "SKU", "Part Name", "Category", "Qty", "Unit",
                       "Unit Cost", "Selling Price", "Disc %", "Disc Amt",
                       "Subtotal", "Total Amount", "Gross Profit", "Margin %",
                       "Reason", "Job Ref", "User"]
            data = self.sales_detail(date_from=date_from, date_to=date_to)

            def row_values(d):
                return [d["issued_at"], d["sku"], d["name"], d["category"],
                        d["quantity"], d["unit"], d["unit_cost"], d["selling_price"],
                        d["discount_pct"], d["discount_amount"], d["subtotal"],
                        d["total_amount"], d["gross_profit"], d["margin_pct"],
                        d["reason"], d["job_ref"], d["issued_by"]]
            num_cols = {7, 8, 10, 11, 12, 13}

        elif report_type == "audit":
            ws.title = "Audit Log"
            headers = ["Timestamp", "Username",
                       "Action", "Module", "Description"]
            data = self.audit_log(date_from=date_from, date_to=date_to)

            def row_values(d):
                return [d.get("timestamp", ""), d.get("username", ""), d.get("action", ""),
                        d.get("module", ""), d.get("description", "")]
            num_cols = set()  # No numeric currency columns needed here

        else:
            raise ValueError(f"Unknown report type: {report_type}")

        # ── Write header row ──────────────────────────────────────────
        ws.row_dimensions[1].height = 32
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = Border(
                bottom=Side(style="medium", color="FFFFFF"),
                right=Side(style="thin", color="2A4A82"))

        # ── Write data rows ───────────────────────────────────────────
        for ri, d in enumerate(data, 2):
            fill = alt_fill if ri % 2 == 0 else PatternFill(
                "solid", fgColor="FFFFFF")
            for ci, val in enumerate(row_values(d), 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.border = cell_border
                if ci in num_cols and isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = num_align
                elif isinstance(val, int):
                    cell.alignment = ctr_align

        # ── Totals rows (sales report) — TOTAL, LABOR, NET TOTAL ──────
        if report_type == "sales" and data:
            sum_total = round(sum(d.get("total_amount", 0) or 0 for d in data), 2)
            sum_profit = round(sum(d.get("gross_profit", 0) or 0 for d in data), 2)

            # Labor is a sale-level fee (not a line item) — extra revenue & profit.
            lwhere, lparams = "WHERE 1=1", {}
            if date_from:
                lwhere += " AND DATE(sale_date) >= :df"
                lparams["df"] = date_from
            if date_to:
                lwhere += " AND DATE(sale_date) <= :dt"
                lparams["dt"] = date_to
            labor = round(self.db.execute(text(
                f"SELECT COALESCE(SUM(labor_amount),0) FROM sales {lwhere}"),
                lparams).scalar() or 0, 2)

            bold = Font(bold=True, color="1D3461")
            top = Side(style="medium", color="1D3461")

            def _totrow(rownum, cells, line=False):
                for ci in range(1, len(headers) + 1):
                    val = cells.get(ci, "")
                    cell = ws.cell(row=rownum, column=ci, value=val)
                    cell.font = bold
                    if line:
                        cell.border = Border(top=top)
                    if ci in num_cols and isinstance(val, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = num_align
                    elif ci == 5:
                        cell.alignment = ctr_align

            tr = len(data) + 2
            _totrow(tr, {
                1:  "TOTAL",
                5:  sum(d.get("quantity", 0) or 0 for d in data),
                10: round(sum(d.get("discount_amount", 0) or 0 for d in data), 2),
                11: round(sum(d.get("subtotal", 0) or 0 for d in data), 2),
                12: sum_total,
                13: sum_profit,
            }, line=True)
            _totrow(tr + 1, {1: "FEES", 12: labor, 13: labor})
            _totrow(tr + 2, {
                1:  "NET TOTAL (incl. labor)",
                12: round(sum_total + labor, 2),
                13: round(sum_profit + labor, 2),
            }, line=True)

        # ── Autofit column widths ─────────────────────────────────────
        for ci in range(1, len(headers) + 1):
            col_letter = get_column_letter(ci)
            max_len = len(str(headers[ci-1]))
            for ri in range(2, len(data) + 2):
                val = ws.cell(row=ri, column=ci).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            # Cap width, add padding
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(filename)
        return filename
